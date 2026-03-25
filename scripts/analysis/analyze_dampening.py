"""analyze_dampening.py — Test dampening factors for opposing pitcher quality adjustments.

Compares adjusted wOBA (dampened by opposing team SIERA ratio) against actual wOBA
across multiple dampening levels and SIERA ratio buckets to find the optimal dampening
factor for pitcher quality adjustments.
"""

import argparse
import logging
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from scripts.analysis.utils import (
    HEADER_FONT,
    PROJECT_ROOT,
    add_conditional_min_max,
    auto_width,
    get_db_connection,
    style_header_row,
)

logger = logging.getLogger(__name__)

DAMPENING_LEVELS = [0.40, 0.45, 0.50, 0.55, 0.60, 0.65, 0.70]
SIERA_BUCKETS = [
    (0.70, 0.80),
    (0.80, 0.90),
    (0.90, 0.95),
    (0.95, 1.05),
    (1.05, 1.10),
    (1.10, 1.20),
    (1.20, 1.30),
]
SEASONS = list(range(2019, 2026))


def load_data(db_path: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Load batting and pitching season data from the backtest database."""
    conn = get_db_connection(db_path)

    batting = pd.read_sql_query(
        """
        SELECT fangraphs_id, season, name, team, PA, wOBA, HR, R, AB
        FROM batting_season
        WHERE season BETWEEN 2019 AND 2025
          AND PA >= 200
          AND wOBA IS NOT NULL
        """,
        conn,
    )

    pitching = pd.read_sql_query(
        """
        SELECT season, team, IP, SIERA
        FROM pitching_season
        WHERE season BETWEEN 2019 AND 2025
          AND IP IS NOT NULL
          AND SIERA IS NOT NULL
        """,
        conn,
    )

    conn.close()
    logger.info("Loaded %d batting rows, %d pitching rows", len(batting), len(pitching))
    return batting, pitching


def run_analysis(
    batting: pd.DataFrame, pitching: pd.DataFrame
) -> dict[str, pd.DataFrame]:
    """Run dampening analysis across all levels and SIERA buckets."""
    # Compute team-level aggregate SIERA (IP-weighted)
    pitching = pitching.dropna(subset=["SIERA", "IP"])
    pitching["weighted_SIERA"] = pitching["SIERA"] * pitching["IP"]
    team_siera = (
        pitching.groupby(["season", "team"])
        .agg(total_IP=("IP", "sum"), weighted_sum=("weighted_SIERA", "sum"))
        .reset_index()
    )
    team_siera["team_SIERA"] = team_siera["weighted_sum"] / team_siera["total_IP"]

    # League average SIERA per season
    league_avg = (
        team_siera.groupby("season")["team_SIERA"].mean().reset_index()
    )
    league_avg.columns = ["season", "league_avg_SIERA"]

    team_siera = team_siera.merge(league_avg, on="season")
    team_siera["SIERA_ratio"] = team_siera["team_SIERA"] / team_siera["league_avg_SIERA"]

    # Assign buckets
    def assign_bucket(ratio: float) -> str:
        for lo, hi in SIERA_BUCKETS:
            if lo <= ratio < hi:
                return f"{lo:.2f}-{hi:.2f}"
        return "out_of_range"

    team_siera["bucket"] = team_siera["SIERA_ratio"].apply(assign_bucket)

    # For each hitter, assume they face a mix of opponents. As a proxy,
    # use the league-average opposing SIERA ratio for their team's opponents.
    # Simplification: use team SIERA as the opposing quality proxy.
    # In reality we'd want schedule-weighted opponent SIERA, but with season
    # data this is a reasonable approximation.
    batting = batting.merge(
        team_siera[["season", "team", "SIERA_ratio", "bucket", "league_avg_SIERA"]],
        on=["season", "team"],
        how="inner",
    )

    # Build raw data with adjusted wOBA for each dampening level
    raw_records = []
    for _, row in batting.iterrows():
        record = {
            "season": int(row["season"]),
            "name": row["name"],
            "team": row["team"],
            "PA": int(row["PA"]),
            "actual_wOBA": round(float(row["wOBA"]), 3),
            "SIERA_ratio": round(float(row["SIERA_ratio"]), 4),
            "bucket": row["bucket"],
        }
        for d in DAMPENING_LEVELS:
            adj = row["wOBA"] * (1 + (row["SIERA_ratio"] - 1) * d)
            record[f"adj_wOBA_{d:.2f}"] = round(float(adj), 4)
        raw_records.append(record)

    raw_df = pd.DataFrame(raw_records)
    raw_df = raw_df[raw_df["bucket"] != "out_of_range"].copy()

    # RMSE by dampening level
    rmse_records = []
    for d in DAMPENING_LEVELS:
        col = f"adj_wOBA_{d:.2f}"
        valid = raw_df.dropna(subset=["actual_wOBA", col])
        if len(valid) > 0:
            rmse = np.sqrt(np.mean((valid["actual_wOBA"] - valid[col]) ** 2))
        else:
            rmse = np.nan
        rmse_records.append({"dampening": d, "RMSE": rmse, "n": len(valid)})
    rmse_df = pd.DataFrame(rmse_records)

    # RMSE by SIERA bucket and dampening level
    bucket_records = []
    for bucket_label in [f"{lo:.2f}-{hi:.2f}" for lo, hi in SIERA_BUCKETS]:
        subset = raw_df[raw_df["bucket"] == bucket_label]
        rec = {"bucket": bucket_label, "n": len(subset)}
        for d in DAMPENING_LEVELS:
            col = f"adj_wOBA_{d:.2f}"
            valid = subset.dropna(subset=["actual_wOBA", col])
            if len(valid) > 0:
                rec[f"RMSE_{d:.2f}"] = np.sqrt(
                    np.mean((valid["actual_wOBA"] - valid[col]) ** 2)
                )
            else:
                rec[f"RMSE_{d:.2f}"] = np.nan
        bucket_records.append(rec)
    bucket_df = pd.DataFrame(bucket_records)

    return {"raw": raw_df, "rmse": rmse_df, "bucket": bucket_df}


def write_excel(results: dict[str, pd.DataFrame], output_path: str) -> None:
    """Write analysis results to formatted Excel workbook with formulas."""
    wb = Workbook()

    # ── Sheet 1: Raw Data ──
    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    headers = [
        "Season", "Name", "Team", "PA", "Actual wOBA", "SIERA Ratio", "Bucket",
    ] + [f"Adj wOBA (d={d:.2f})" for d in DAMPENING_LEVELS]

    for c, h in enumerate(headers, 1):
        ws_raw.cell(row=1, column=c, value=h)
    style_header_row(ws_raw, len(headers))

    raw_df = results["raw"]
    for r_idx, (_, row) in enumerate(raw_df.iterrows(), 2):
        ws_raw.cell(row=r_idx, column=1, value=row["season"])
        ws_raw.cell(row=r_idx, column=2, value=row["name"])
        ws_raw.cell(row=r_idx, column=3, value=row["team"])
        ws_raw.cell(row=r_idx, column=4, value=row["PA"])
        ws_raw.cell(row=r_idx, column=5, value=row["actual_wOBA"])
        ws_raw.cell(row=r_idx, column=5).number_format = "0.000"
        ws_raw.cell(row=r_idx, column=6, value=row["SIERA_ratio"])
        ws_raw.cell(row=r_idx, column=6).number_format = "0.0000"
        ws_raw.cell(row=r_idx, column=7, value=row["bucket"])
        for d_idx, d in enumerate(DAMPENING_LEVELS):
            cell = ws_raw.cell(
                row=r_idx, column=8 + d_idx, value=row[f"adj_wOBA_{d:.2f}"]
            )
            cell.number_format = "0.0000"

    ws_raw.cell(row=1, column=5).comment = Comment(
        "Player's actual season wOBA from FanGraphs batting_season table.",
        "Analysis Script",
    )
    ws_raw.cell(row=1, column=6).comment = Comment(
        "SIERA ratio = team SIERA / league avg SIERA.\n"
        ">1.0 means worse pitching staff (hitter-friendly),\n"
        "<1.0 means better pitching staff (pitcher-friendly).",
        "Analysis Script",
    )

    auto_width(ws_raw)

    # ── Sheet 2: RMSE by Dampening ──
    ws_rmse = wb.create_sheet("RMSE by Dampening")

    last_data_row = len(raw_df) + 1  # row count in Raw Data sheet

    headers_rmse = ["Dampening Level", "RMSE (Formula)", "RMSE (Computed)", "N"]
    for c, h in enumerate(headers_rmse, 1):
        ws_rmse.cell(row=1, column=c, value=h)
    style_header_row(ws_rmse, len(headers_rmse))

    for r_idx, d in enumerate(DAMPENING_LEVELS, 2):
        ws_rmse.cell(row=r_idx, column=1, value=d)
        ws_rmse.cell(row=r_idx, column=1).number_format = "0.00"

        # Excel formula referencing Raw Data sheet
        actual_col = "E"  # actual wOBA
        adj_col = get_column_letter(8 + DAMPENING_LEVELS.index(d))  # adjusted wOBA col

        formula = (
            f"=SQRT(SUMPRODUCT(('Raw Data'!{actual_col}2:{actual_col}{last_data_row}"
            f"-'Raw Data'!{adj_col}2:{adj_col}{last_data_row})^2)"
            f"/COUNTA('Raw Data'!{actual_col}2:{actual_col}{last_data_row}))"
        )
        cell_f = ws_rmse.cell(row=r_idx, column=2, value=formula)
        cell_f.number_format = "0.000000"

        # Static computed value for reference
        rmse_row = results["rmse"][results["rmse"]["dampening"] == d]
        if not rmse_row.empty:
            ws_rmse.cell(row=r_idx, column=3, value=round(rmse_row.iloc[0]["RMSE"], 6))
            ws_rmse.cell(row=r_idx, column=3).number_format = "0.000000"
            ws_rmse.cell(row=r_idx, column=4, value=int(rmse_row.iloc[0]["n"]))

    ws_rmse.cell(row=1, column=2).comment = Comment(
        "RMSE computed via Excel formula: SQRT(SUMPRODUCT((actual-adjusted)^2)/N).\n"
        "References the Raw Data sheet directly so values update if data changes.",
        "Analysis Script",
    )

    # Conditional formatting: green for lowest RMSE, red for highest
    rmse_range = f"B2:B{1 + len(DAMPENING_LEVELS)}"
    add_conditional_min_max(ws_rmse, rmse_range, green_best="min")

    auto_width(ws_rmse)

    # ── Sheet 3: By SIERA Bucket ──
    ws_bucket = wb.create_sheet("By SIERA Bucket")

    bucket_headers = ["SIERA Bucket", "N"] + [
        f"RMSE (d={d:.2f})" for d in DAMPENING_LEVELS
    ]
    for c, h in enumerate(bucket_headers, 1):
        ws_bucket.cell(row=1, column=c, value=h)
    style_header_row(ws_bucket, len(bucket_headers))

    bucket_df = results["bucket"]
    for r_idx, (_, row) in enumerate(bucket_df.iterrows(), 2):
        ws_bucket.cell(row=r_idx, column=1, value=row["bucket"])
        ws_bucket.cell(row=r_idx, column=2, value=int(row["n"]) if not pd.isna(row["n"]) else 0)
        for d_idx, d in enumerate(DAMPENING_LEVELS):
            val = row.get(f"RMSE_{d:.2f}", np.nan)
            cell = ws_bucket.cell(
                row=r_idx, column=3 + d_idx,
                value=round(float(val), 6) if not pd.isna(val) else None,
            )
            cell.number_format = "0.000000"

    # Highlight min RMSE per row
    for r_idx in range(2, 2 + len(bucket_df)):
        start_col = get_column_letter(3)
        end_col = get_column_letter(3 + len(DAMPENING_LEVELS) - 1)
        rng = f"{start_col}{r_idx}:{end_col}{r_idx}"
        add_conditional_min_max(ws_bucket, rng, green_best="min")

    ws_bucket.cell(row=1, column=1).comment = Comment(
        "SIERA ratio buckets represent opposing pitching quality.\n"
        "Lower ratios = better opposing pitching, higher = weaker.",
        "Analysis Script",
    )

    auto_width(ws_bucket)

    # ── Sheet 4: Recommendation ──
    ws_rec = wb.create_sheet("Recommendation")

    ws_rec.cell(row=1, column=1, value="Optimal Dampening Analysis").font = Font(
        bold=True, size=14
    )

    ws_rec.cell(row=3, column=1, value="Optimal Dampening Level:")
    ws_rec.cell(row=3, column=1).font = HEADER_FONT

    # Formula to find the dampening with the minimum RMSE
    rmse_last = 1 + len(DAMPENING_LEVELS)
    ws_rec.cell(
        row=3,
        column=2,
        value=f"=INDEX('RMSE by Dampening'!A2:A{rmse_last},"
        f"MATCH(MIN('RMSE by Dampening'!B2:B{rmse_last}),"
        f"'RMSE by Dampening'!B2:B{rmse_last},0))",
    )
    ws_rec.cell(row=3, column=2).number_format = "0.00"
    ws_rec.cell(row=3, column=2).comment = Comment(
        "This formula finds the dampening level with the lowest RMSE\n"
        "from the 'RMSE by Dampening' sheet using INDEX/MATCH.\n\n"
        "Dampening controls how much opposing pitcher quality affects\n"
        "projected wOBA. Lower dampening = less adjustment,\n"
        "higher dampening = more aggressive adjustment.",
        "Analysis Script",
    )

    ws_rec.cell(row=4, column=1, value="Minimum RMSE:")
    ws_rec.cell(row=4, column=1).font = HEADER_FONT
    ws_rec.cell(
        row=4,
        column=2,
        value=f"=MIN('RMSE by Dampening'!B2:B{rmse_last})",
    )
    ws_rec.cell(row=4, column=2).number_format = "0.000000"

    ws_rec.cell(row=6, column=1, value="Methodology:").font = HEADER_FONT
    ws_rec.cell(row=7, column=1, value=(
        "For each hitter-season (2019-2025, 200+ PA), we compute adjusted wOBA as:\n"
        "  adjusted_wOBA = actual_wOBA * (1 + (SIERA_ratio - 1) * dampening)\n"
        "where SIERA_ratio = team_SIERA / league_avg_SIERA.\n\n"
        "We test dampening levels from 0.40 to 0.70 and compute RMSE of\n"
        "adjusted vs actual wOBA. The optimal dampening minimizes RMSE."
    ))
    ws_rec.cell(row=7, column=1).alignment = Alignment(wrap_text=True)
    ws_rec.column_dimensions["A"].width = 30
    ws_rec.column_dimensions["B"].width = 20

    auto_width(ws_rec)

    wb.save(output_path)
    logger.info("Saved workbook to %s", output_path)


def main() -> None:
    """Entry point for dampening analysis."""
    parser = argparse.ArgumentParser(
        description="Analyze dampening factors for pitcher quality adjustments"
    )
    parser.add_argument(
        "--db-path",
        default=str(PROJECT_ROOT / "backtest_data.sqlite"),
        help="Path to backtest SQLite database (default: backtest_data.sqlite)",
    )
    parser.add_argument(
        "--output-dir",
        default=str(PROJECT_ROOT / "analysis"),
        help="Directory for Excel output",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(name)s %(levelname)s %(message)s",
    )

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    batting, pitching = load_data(args.db_path)

    if batting.empty or pitching.empty:
        logger.warning("No data found. Generating empty workbook.")
        wb = Workbook()
        wb.active.title = "No Data"
        wb.active.cell(row=1, column=1, value="No data found in database.")
        wb.save(str(output_dir / "dampening_analysis.xlsx"))
        return

    results = run_analysis(batting, pitching)
    write_excel(results, str(output_dir / "dampening_analysis.xlsx"))
    logger.info("Dampening analysis complete.")


if __name__ == "__main__":
    main()
