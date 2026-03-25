"""analyze_park_factors.py — Test park factor strength multipliers.

Evaluates how aggressively park factors should be applied to HR and R rate
projections by testing multiple strength multipliers against actual production.
"""

import argparse
import logging
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from scripts.analysis.utils import (
    HEADER_FONT,
    PROJECT_ROOT,
    add_conditional_min_max,
    auto_width,
    get_db_connection,
    style_header_row,
)

logger = logging.getLogger(__name__)

STRENGTH_LEVELS = [0.60, 0.65, 0.70, 0.75, 0.80, 0.85, 0.90, 0.95, 1.00]

# Hardcoded park factors (HR factor, basic/runs factor) relative to 100 = neutral
# Sources: FanGraphs park factors, multi-year averages
PARK_FACTORS: dict[str, dict[str, int]] = {
    "COL": {"HR": 116, "basic": 118},
    "CIN": {"HR": 112, "basic": 106},
    "TEX": {"HR": 108, "basic": 104},
    "PHI": {"HR": 108, "basic": 103},
    "MIL": {"HR": 107, "basic": 103},
    "TOR": {"HR": 107, "basic": 104},
    "CHC": {"HR": 106, "basic": 104},
    "BOS": {"HR": 104, "basic": 105},
    "BAL": {"HR": 106, "basic": 103},
    "ATL": {"HR": 104, "basic": 102},
    "NYY": {"HR": 110, "basic": 103},
    "ARI": {"HR": 103, "basic": 103},
    "MIN": {"HR": 103, "basic": 101},
    "LAA": {"HR": 100, "basic": 100},
    "CHW": {"HR": 100, "basic": 100},
    "CLE": {"HR": 98, "basic": 99},
    "WSN": {"HR": 99, "basic": 100},
    "DET": {"HR": 97, "basic": 99},
    "HOU": {"HR": 100, "basic": 101},
    "PIT": {"HR": 96, "basic": 98},
    "KCR": {"HR": 95, "basic": 99},
    "SEA": {"HR": 96, "basic": 97},
    "STL": {"HR": 97, "basic": 98},
    "SDP": {"HR": 94, "basic": 96},
    "LAD": {"HR": 95, "basic": 97},
    "TBR": {"HR": 93, "basic": 96},
    "SFG": {"HR": 90, "basic": 95},
    "NYM": {"HR": 93, "basic": 97},
    "MIA": {"HR": 90, "basic": 94},
    "OAK": {"HR": 91, "basic": 95},
}

# Aliases for team abbreviations
TEAM_ALIASES: dict[str, str] = {
    "CWS": "CHW",
    "SD": "SDP",
    "SF": "SFG",
    "TB": "TBR",
    "KC": "KCR",
    "WSH": "WSN",
    "WAS": "WSN",
    "ANA": "LAA",
}

def _normalize_team(team: str) -> str:
    t = str(team).strip().upper()
    return TEAM_ALIASES.get(t, t)


def load_data(db_path: str) -> pd.DataFrame:
    """Load batting season data."""
    conn = get_db_connection(db_path)

    batting = pd.read_sql_query(
        """
        SELECT fangraphs_id, season, name, team, PA, AB, HR, R
        FROM batting_season
        WHERE season BETWEEN 2019 AND 2025
          AND PA >= 200
          AND AB > 0
        """,
        conn,
    )
    conn.close()
    logger.info("Loaded %d batting rows", len(batting))
    return batting


def _get_park_factor(team: str, stat: str) -> int:
    """Look up park factor for a team; return 100 (neutral) if unknown."""
    t = _normalize_team(team)
    factors = PARK_FACTORS.get(t, {"HR": 100, "basic": 100})
    return factors.get(stat, 100)


def run_analysis(batting: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Test park factor strength multipliers against actual HR and R rates."""
    batting = batting.copy()
    batting["team_norm"] = batting["team"].apply(_normalize_team)
    batting["HR_rate"] = batting["HR"] / batting["AB"]
    batting["R_rate"] = batting["R"] / batting["PA"]
    batting["park_HR"] = batting["team_norm"].apply(lambda t: _get_park_factor(t, "HR"))
    batting["park_basic"] = batting["team_norm"].apply(
        lambda t: _get_park_factor(t, "basic")
    )

    # League average rates per season for neutral baseline
    league_avg = (
        batting.groupby("season")
        .agg(lg_HR_rate=("HR_rate", "mean"), lg_R_rate=("R_rate", "mean"))
        .reset_index()
    )
    batting = batting.merge(league_avg, on="season")

    # Raw data with adjusted rates per strength
    raw_records = []
    for _, row in batting.iterrows():
        rec = {
            "season": int(row["season"]),
            "name": row["name"],
            "team": row["team_norm"],
            "PA": int(row["PA"]),
            "actual_HR_rate": row["HR_rate"],
            "actual_R_rate": row["R_rate"],
            "park_HR_factor": row["park_HR"],
            "park_basic_factor": row["park_basic"],
        }
        for s in STRENGTH_LEVELS:
            rec[f"adj_HR_{s:.2f}"] = row["HR_rate"] * (
                1 + (row["park_HR"] / 100 - 1) * s
            )
            rec[f"adj_R_{s:.2f}"] = row["R_rate"] * (
                1 + (row["park_basic"] / 100 - 1) * s
            )
        raw_records.append(rec)

    raw_df = pd.DataFrame(raw_records)

    # RMSE by strength level
    strength_records = []
    for s in STRENGTH_LEVELS:
        valid = raw_df.dropna(subset=["actual_HR_rate", f"adj_HR_{s:.2f}"])
        hr_rmse = np.sqrt(
            np.mean((valid["actual_HR_rate"] - valid[f"adj_HR_{s:.2f}"]) ** 2)
        ) if len(valid) > 0 else np.nan

        valid_r = raw_df.dropna(subset=["actual_R_rate", f"adj_R_{s:.2f}"])
        r_rmse = np.sqrt(
            np.mean((valid_r["actual_R_rate"] - valid_r[f"adj_R_{s:.2f}"]) ** 2)
        ) if len(valid_r) > 0 else np.nan

        strength_records.append({
            "strength": s,
            "HR_RMSE": hr_rmse,
            "R_RMSE": r_rmse,
            "n": len(valid),
        })
    strength_df = pd.DataFrame(strength_records)

    # By park analysis
    park_records = []
    for team in sorted(raw_df["team"].unique()):
        subset = raw_df[raw_df["team"] == team]
        if len(subset) < 5:
            continue
        rec = {
            "team": team,
            "park_HR_factor": subset["park_HR_factor"].iloc[0],
            "park_basic_factor": subset["park_basic_factor"].iloc[0],
            "n": len(subset),
        }
        best_hr_rmse = float("inf")
        best_hr_strength = None
        best_r_rmse = float("inf")
        best_r_strength = None
        for s in STRENGTH_LEVELS:
            hr_rmse = np.sqrt(
                np.mean((subset["actual_HR_rate"] - subset[f"adj_HR_{s:.2f}"]) ** 2)
            )
            r_rmse = np.sqrt(
                np.mean((subset["actual_R_rate"] - subset[f"adj_R_{s:.2f}"]) ** 2)
            )
            rec[f"HR_RMSE_{s:.2f}"] = hr_rmse
            rec[f"R_RMSE_{s:.2f}"] = r_rmse
            if hr_rmse < best_hr_rmse:
                best_hr_rmse = hr_rmse
                best_hr_strength = s
            if r_rmse < best_r_rmse:
                best_r_rmse = r_rmse
                best_r_strength = s
        rec["best_HR_strength"] = best_hr_strength
        rec["best_R_strength"] = best_r_strength
        park_records.append(rec)
    park_df = pd.DataFrame(park_records)

    # Extreme parks
    if not park_df.empty:
        park_df["deviation"] = abs(park_df["park_HR_factor"] - 100) + abs(
            park_df["park_basic_factor"] - 100
        )
        hitter_friendly = (
            park_df[park_df["park_HR_factor"] > 100]
            .nlargest(5, "deviation")
            .copy()
        )
        pitcher_friendly = (
            park_df[park_df["park_HR_factor"] < 100]
            .nlargest(5, "deviation")
            .copy()
        )
    else:
        hitter_friendly = pd.DataFrame()
        pitcher_friendly = pd.DataFrame()

    return {
        "raw": raw_df,
        "strength": strength_df,
        "park": park_df,
        "hitter_friendly": hitter_friendly,
        "pitcher_friendly": pitcher_friendly,
    }


def write_excel(results: dict[str, pd.DataFrame], output_path: str) -> None:
    """Write park factor analysis to formatted Excel."""
    wb = Workbook()

    raw_df = results["raw"]
    strength_df = results["strength"]
    park_df = results["park"]

    # ── Sheet 1: By Strength ──
    ws = wb.active
    ws.title = "By Strength"

    # Write raw data first (hidden, for formula references)
    ws_data = wb.create_sheet("_RawData")
    data_headers = [
        "Season", "Name", "Team", "PA", "Actual HR Rate", "Actual R Rate",
        "Park HR Factor", "Park Basic Factor",
    ]
    for i, s in enumerate(STRENGTH_LEVELS):
        data_headers.append(f"Adj HR {s:.2f}")
    for i, s in enumerate(STRENGTH_LEVELS):
        data_headers.append(f"Adj R {s:.2f}")

    for c, h in enumerate(data_headers, 1):
        ws_data.cell(row=1, column=c, value=h)
    style_header_row(ws_data, len(data_headers))

    for r_idx, (_, row) in enumerate(raw_df.iterrows(), 2):
        ws_data.cell(row=r_idx, column=1, value=row["season"])
        ws_data.cell(row=r_idx, column=2, value=row["name"])
        ws_data.cell(row=r_idx, column=3, value=row["team"])
        ws_data.cell(row=r_idx, column=4, value=row["PA"])
        ws_data.cell(row=r_idx, column=5, value=row["actual_HR_rate"])
        ws_data.cell(row=r_idx, column=5).number_format = "0.0000"
        ws_data.cell(row=r_idx, column=6, value=row["actual_R_rate"])
        ws_data.cell(row=r_idx, column=6).number_format = "0.0000"
        ws_data.cell(row=r_idx, column=7, value=row["park_HR_factor"])
        ws_data.cell(row=r_idx, column=8, value=row["park_basic_factor"])
        for s_idx, s in enumerate(STRENGTH_LEVELS):
            ws_data.cell(
                row=r_idx, column=9 + s_idx, value=row[f"adj_HR_{s:.2f}"]
            ).number_format = "0.0000"
            ws_data.cell(
                row=r_idx,
                column=9 + len(STRENGTH_LEVELS) + s_idx,
                value=row[f"adj_R_{s:.2f}"],
            ).number_format = "0.0000"

    last_row = len(raw_df) + 1

    # By Strength sheet with formulas
    headers = ["Strength", "HR Rate RMSE (Formula)", "R Rate RMSE (Formula)",
               "HR RMSE (Computed)", "R RMSE (Computed)", "N"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, len(headers))

    actual_hr_col = "E"  # actual HR rate in _RawData
    actual_r_col = "F"   # actual R rate in _RawData

    for r_idx, (i, s) in enumerate(enumerate(STRENGTH_LEVELS), 2):
        ws.cell(row=r_idx, column=1, value=s)
        ws.cell(row=r_idx, column=1).number_format = "0.00"

        adj_hr_col = get_column_letter(9 + i)
        adj_r_col = get_column_letter(9 + len(STRENGTH_LEVELS) + i)

        hr_formula = (
            f"=SQRT(SUMPRODUCT(('_RawData'!{actual_hr_col}2:{actual_hr_col}{last_row}"
            f"-'_RawData'!{adj_hr_col}2:{adj_hr_col}{last_row})^2)"
            f"/COUNTA('_RawData'!{actual_hr_col}2:{actual_hr_col}{last_row}))"
        )
        r_formula = (
            f"=SQRT(SUMPRODUCT(('_RawData'!{actual_r_col}2:{actual_r_col}{last_row}"
            f"-'_RawData'!{adj_r_col}2:{adj_r_col}{last_row})^2)"
            f"/COUNTA('_RawData'!{actual_r_col}2:{actual_r_col}{last_row}))"
        )

        ws.cell(row=r_idx, column=2, value=hr_formula).number_format = "0.000000"
        ws.cell(row=r_idx, column=3, value=r_formula).number_format = "0.000000"

        s_row = strength_df[strength_df["strength"] == s]
        if not s_row.empty:
            ws.cell(row=r_idx, column=4, value=round(s_row.iloc[0]["HR_RMSE"], 6))
            ws.cell(row=r_idx, column=4).number_format = "0.000000"
            ws.cell(row=r_idx, column=5, value=round(s_row.iloc[0]["R_RMSE"], 6))
            ws.cell(row=r_idx, column=5).number_format = "0.000000"
            ws.cell(row=r_idx, column=6, value=int(s_row.iloc[0]["n"]))

    # Conditional formatting: green=lowest RMSE, red=highest
    sr = len(STRENGTH_LEVELS)
    for col_letter in ["B", "C"]:
        rng = f"{col_letter}2:{col_letter}{1 + sr}"
        add_conditional_min_max(ws, rng, green_best="min")

    ws.cell(row=1, column=2).comment = Comment(
        "HR rate RMSE: measures prediction error for home run rate\n"
        "after applying park factor at this strength level.\n"
        "Formula: SQRT(SUMPRODUCT((actual-adjusted)^2)/N)",
        "Analysis Script",
    )

    auto_width(ws)

    # ── Sheet 2: By Park ──
    ws_park = wb.create_sheet("By Park")
    park_headers = [
        "Team", "Park HR Factor", "Park Basic Factor", "N",
        "Best HR Strength", "Best R Strength",
    ]
    for c, h in enumerate(park_headers, 1):
        ws_park.cell(row=1, column=c, value=h)
    style_header_row(ws_park, len(park_headers))

    for r_idx, (_, row) in enumerate(park_df.iterrows(), 2):
        ws_park.cell(row=r_idx, column=1, value=row["team"])
        ws_park.cell(row=r_idx, column=2, value=int(row["park_HR_factor"]))
        ws_park.cell(row=r_idx, column=3, value=int(row["park_basic_factor"]))
        ws_park.cell(row=r_idx, column=4, value=int(row["n"]))
        ws_park.cell(row=r_idx, column=5, value=row.get("best_HR_strength"))
        ws_park.cell(row=r_idx, column=5).number_format = "0.00"
        ws_park.cell(row=r_idx, column=6, value=row.get("best_R_strength"))
        ws_park.cell(row=r_idx, column=6).number_format = "0.00"

    ws_park.cell(row=1, column=5).comment = Comment(
        "The strength multiplier that produced the lowest RMSE for HR rate\n"
        "at this specific park. Different parks may benefit from different strengths.",
        "Analysis Script",
    )

    auto_width(ws_park)

    # ── Sheet 3: Extreme Parks ──
    ws_ext = wb.create_sheet("Extreme Parks")
    ws_ext.cell(row=1, column=1, value="Top 5 Hitter-Friendly Parks").font = Font(
        bold=True, size=13
    )
    ext_headers = ["Team", "HR Factor", "Basic Factor", "Best HR Strength", "Best R Strength"]
    for c, h in enumerate(ext_headers, 1):
        ws_ext.cell(row=2, column=c, value=h)
    style_header_row(ws_ext, len(ext_headers), row=2)

    hf = results["hitter_friendly"]
    for r_idx, (_, row) in enumerate(hf.iterrows(), 3):
        ws_ext.cell(row=r_idx, column=1, value=row["team"])
        ws_ext.cell(row=r_idx, column=2, value=int(row["park_HR_factor"]))
        ws_ext.cell(row=r_idx, column=3, value=int(row["park_basic_factor"]))
        ws_ext.cell(row=r_idx, column=4, value=row.get("best_HR_strength"))
        ws_ext.cell(row=r_idx, column=5, value=row.get("best_R_strength"))

    offset = 3 + len(hf) + 1
    ws_ext.cell(row=offset, column=1, value="Top 5 Pitcher-Friendly Parks").font = Font(
        bold=True, size=13
    )
    for c, h in enumerate(ext_headers, 1):
        ws_ext.cell(row=offset + 1, column=c, value=h)
    style_header_row(ws_ext, len(ext_headers), row=offset + 1)

    pf = results["pitcher_friendly"]
    for r_idx, (_, row) in enumerate(pf.iterrows(), offset + 2):
        ws_ext.cell(row=r_idx, column=1, value=row["team"])
        ws_ext.cell(row=r_idx, column=2, value=int(row["park_HR_factor"]))
        ws_ext.cell(row=r_idx, column=3, value=int(row["park_basic_factor"]))
        ws_ext.cell(row=r_idx, column=4, value=row.get("best_HR_strength"))
        ws_ext.cell(row=r_idx, column=5, value=row.get("best_R_strength"))

    auto_width(ws_ext)

    # ── Sheet 4: Recommendation ──
    ws_rec = wb.create_sheet("Recommendation")
    ws_rec.cell(row=1, column=1, value="Park Factor Strength Recommendation").font = Font(
        bold=True, size=14
    )

    sr = len(STRENGTH_LEVELS)
    ws_rec.cell(row=3, column=1, value="Optimal HR Strength:").font = HEADER_FONT
    ws_rec.cell(
        row=3, column=2,
        value=f"=INDEX('By Strength'!A2:A{1+sr},"
        f"MATCH(MIN('By Strength'!B2:B{1+sr}),'By Strength'!B2:B{1+sr},0))",
    )
    ws_rec.cell(row=3, column=2).number_format = "0.00"
    ws_rec.cell(row=3, column=2).comment = Comment(
        "Strength multiplier with lowest HR rate RMSE.\n"
        "Applied as: adj_HR_rate = HR_rate * (1 + (park_factor/100 - 1) * strength)",
        "Analysis Script",
    )

    ws_rec.cell(row=4, column=1, value="Optimal R Strength:").font = HEADER_FONT
    ws_rec.cell(
        row=4, column=2,
        value=f"=INDEX('By Strength'!A2:A{1+sr},"
        f"MATCH(MIN('By Strength'!C2:C{1+sr}),'By Strength'!C2:C{1+sr},0))",
    )
    ws_rec.cell(row=4, column=2).number_format = "0.00"

    ws_rec.cell(row=6, column=1, value="Notes:").font = HEADER_FONT
    ws_rec.cell(row=7, column=1, value=(
        "Park factor strength controls how aggressively we adjust for park effects.\n"
        "A strength of 1.00 applies the full park factor; 0.50 applies half.\n"
        "Lower strengths may be optimal because park factors already contain noise\n"
        "and year-to-year variance."
    )).alignment = Alignment(wrap_text=True)
    ws_rec.column_dimensions["A"].width = 30
    ws_rec.column_dimensions["B"].width = 20

    wb.save(output_path)
    logger.info("Saved workbook to %s", output_path)


def main() -> None:
    """Entry point for park factor analysis."""
    parser = argparse.ArgumentParser(
        description="Analyze park factor strength multipliers"
    )
    parser.add_argument(
        "--db-path",
        default=str(PROJECT_ROOT / "backtest_data.sqlite"),
        help="Path to backtest SQLite database",
    )
    parser.add_argument(
        "--output-dir",
        default=str(PROJECT_ROOT / "analysis"),
        help="Directory for Excel output",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(name)s %(levelname)s %(message)s",
    )

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    batting = load_data(args.db_path)

    if batting.empty:
        logger.warning("No data found. Generating empty workbook.")
        wb = Workbook()
        wb.active.title = "No Data"
        wb.active.cell(row=1, column=1, value="No data found in database.")
        wb.save(str(output_dir / "park_factors_analysis.xlsx"))
        return

    results = run_analysis(batting)
    write_excel(results, str(output_dir / "park_factors_analysis.xlsx"))
    logger.info("Park factor analysis complete.")


if __name__ == "__main__":
    main()
