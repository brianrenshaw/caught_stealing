"""analyze_xwoba_regression.py — Test whether Statcast xwOBA predicts future wOBA better than traditional wOBA.

Uses prior-season xwOBA vs wOBA to predict current-season wOBA, computing R-squared
and RMSE for each predictor and various blend ratios.
"""

import argparse
import logging
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from scripts.analysis.utils import (
    HEADER_FONT,
    PROJECT_ROOT,
    add_conditional_min_max,
    auto_width,
    get_db_connection,
    load_player_id_crosswalk,
    style_header_row,
)

logger = logging.getLogger(__name__)

BLEND_RATIOS = [
    (1.0, 0.0, "wOBA only"),
    (0.0, 1.0, "xwOBA only"),
    (0.50, 0.50, "50/50"),
    (0.40, 0.60, "40/60 (wOBA/xwOBA)"),
    (0.30, 0.70, "30/70 (wOBA/xwOBA)"),
]

def _r_squared(y_true: np.ndarray, y_pred: np.ndarray) -> float:
    """Compute R-squared."""
    ss_res = np.sum((y_true - y_pred) ** 2)
    ss_tot = np.sum((y_true - np.mean(y_true)) ** 2)
    if ss_tot == 0:
        return np.nan
    return 1 - ss_res / ss_tot


def _rmse(y_true: np.ndarray, y_pred: np.ndarray) -> float:
    """Compute RMSE."""
    return float(np.sqrt(np.mean((y_true - y_pred) ** 2)))


def load_data(db_path: str) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Load batting, statcast, and player_ids for crosswalk."""
    conn = get_db_connection(db_path)

    batting = pd.read_sql_query(
        """
        SELECT fangraphs_id, season, name, team, PA, wOBA
        FROM batting_season
        WHERE season BETWEEN 2015 AND 2025
          AND PA >= 300
          AND wOBA IS NOT NULL
        """,
        conn,
    )

    statcast = pd.read_sql_query(
        """
        SELECT mlbam_id, season, name, PA, xwoba
        FROM statcast_season
        WHERE season BETWEEN 2015 AND 2025
          AND PA >= 300
          AND xwoba IS NOT NULL
        """,
        conn,
    )

    conn.close()

    player_ids = load_player_id_crosswalk(db_path)

    logger.info(
        "Loaded %d batting, %d statcast, %d id mappings",
        len(batting), len(statcast), len(player_ids),
    )
    return batting, statcast, player_ids


def run_analysis(
    batting: pd.DataFrame,
    statcast: pd.DataFrame,
    player_ids: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    """Build prior-to-current season pairs and test predictors."""
    # Ensure consistent ID types for merging
    player_ids["fangraphs_id"] = player_ids["fangraphs_id"].astype(str)
    player_ids["mlbam_id"] = player_ids["mlbam_id"].astype(str)
    batting["fangraphs_id"] = batting["fangraphs_id"].astype(str)
    statcast["mlbam_id"] = statcast["mlbam_id"].astype(str)

    # Merge statcast with player_ids to get fangraphs_id
    statcast_with_fg = statcast.merge(
        player_ids[["fangraphs_id", "mlbam_id"]].drop_duplicates(),
        on="mlbam_id",
        how="inner",
    )

    # Build pairs: prior season stats -> current season wOBA
    batting_prior = batting.rename(
        columns={"wOBA": "prior_wOBA", "PA": "prior_PA", "season": "prior_season"}
    )
    batting_current = batting.rename(
        columns={"wOBA": "current_wOBA", "PA": "current_PA", "season": "current_season"}
    )
    statcast_prior = statcast_with_fg.rename(
        columns={"xwoba": "prior_xwOBA", "PA": "prior_sc_PA", "season": "prior_season"}
    )

    # Create year-over-year pairs
    pairs = []
    for season in range(2016, 2026):
        prior = batting_prior[batting_prior["prior_season"] == season - 1][
            ["fangraphs_id", "prior_wOBA", "prior_PA", "prior_season"]
        ].drop_duplicates(subset=["fangraphs_id"])

        current = batting_current[batting_current["current_season"] == season][
            ["fangraphs_id", "current_wOBA", "current_PA", "current_season", "name"]
        ].drop_duplicates(subset=["fangraphs_id"])

        sc_prior = statcast_prior[statcast_prior["prior_season"] == season - 1][
            ["fangraphs_id", "prior_xwOBA", "prior_sc_PA"]
        ].drop_duplicates(subset=["fangraphs_id"])

        merged = prior.merge(current, on="fangraphs_id", how="inner")
        merged = merged.merge(sc_prior, on="fangraphs_id", how="inner")
        pairs.append(merged)

    if not pairs:
        return {"scatter": pd.DataFrame(), "comparison": pd.DataFrame(),
                "by_season": pd.DataFrame()}

    all_pairs = pd.concat(pairs, ignore_index=True)
    all_pairs = all_pairs.dropna(subset=["prior_wOBA", "prior_xwOBA", "current_wOBA"])

    logger.info("Built %d year-over-year pairs", len(all_pairs))

    # Compute blended predictors
    for w_woba, w_xwoba, label in BLEND_RATIOS:
        col_name = f"blend_{w_woba:.2f}_{w_xwoba:.2f}"
        all_pairs[col_name] = (
            w_woba * all_pairs["prior_wOBA"] + w_xwoba * all_pairs["prior_xwOBA"]
        )

    # Overall comparison
    comparison_records = []
    for w_woba, w_xwoba, label in BLEND_RATIOS:
        col_name = f"blend_{w_woba:.2f}_{w_xwoba:.2f}"
        valid = all_pairs.dropna(subset=["current_wOBA", col_name])
        y_true = valid["current_wOBA"].values
        y_pred = valid[col_name].values
        comparison_records.append({
            "Predictor": label,
            "wOBA Weight": w_woba,
            "xwOBA Weight": w_xwoba,
            "R_squared": _r_squared(y_true, y_pred),
            "RMSE": _rmse(y_true, y_pred),
            "N": len(valid),
        })
    comparison_df = pd.DataFrame(comparison_records)

    # By season breakdown
    season_records = []
    for season in sorted(all_pairs["current_season"].unique()):
        subset = all_pairs[all_pairs["current_season"] == season]
        for w_woba, w_xwoba, label in BLEND_RATIOS:
            col_name = f"blend_{w_woba:.2f}_{w_xwoba:.2f}"
            valid = subset.dropna(subset=["current_wOBA", col_name])
            if len(valid) < 10:
                continue
            y_true = valid["current_wOBA"].values
            y_pred = valid[col_name].values
            season_records.append({
                "Season": int(season),
                "Predictor": label,
                "R_squared": _r_squared(y_true, y_pred),
                "RMSE": _rmse(y_true, y_pred),
                "N": len(valid),
            })
    season_df = pd.DataFrame(season_records)

    # Scatter data (for external plotting)
    scatter_df = all_pairs[[
        "name", "prior_season", "current_season",
        "prior_wOBA", "prior_xwOBA", "current_wOBA",
    ]].copy()
    scatter_df.columns = [
        "Name", "Prior Season", "Current Season",
        "Prior wOBA", "Prior xwOBA", "Current wOBA",
    ]

    return {
        "scatter": scatter_df,
        "comparison": comparison_df,
        "by_season": season_df,
    }


def write_excel(results: dict[str, pd.DataFrame], output_path: str) -> None:
    """Write xwOBA regression analysis to formatted Excel."""
    wb = Workbook()

    scatter_df = results["scatter"]
    comparison_df = results["comparison"]
    season_df = results["by_season"]

    # ── Sheet 1: Scatter Data (write first for formula references) ──
    ws_scatter = wb.active
    ws_scatter.title = "Scatter Data"

    scatter_headers = list(scatter_df.columns) if not scatter_df.empty else [
        "Name", "Prior Season", "Current Season", "Prior wOBA", "Prior xwOBA", "Current wOBA"
    ]
    for c, h in enumerate(scatter_headers, 1):
        ws_scatter.cell(row=1, column=c, value=h)
    style_header_row(ws_scatter, len(scatter_headers))

    for r_idx, (_, row) in enumerate(scatter_df.iterrows(), 2):
        for c_idx, col in enumerate(scatter_df.columns, 1):
            cell = ws_scatter.cell(row=r_idx, column=c_idx, value=row[col])
            if "wOBA" in col or "xwOBA" in col:
                cell.number_format = "0.000"

    last_scatter = len(scatter_df) + 1

    ws_scatter.cell(row=1, column=4).comment = Comment(
        "Prior season traditional wOBA from FanGraphs.\n"
        "This is the 'actual outcome' measure based on real batting results.",
        "Analysis Script",
    )
    ws_scatter.cell(row=1, column=5).comment = Comment(
        "Prior season expected wOBA from Statcast.\n"
        "Based on exit velocity, launch angle, and sprint speed.\n"
        "Strips out defense and luck to measure quality of contact.",
        "Analysis Script",
    )

    auto_width(ws_scatter)

    # ── Sheet 2: R-Squared Comparison ──
    ws_r2 = wb.create_sheet("R-Squared Comparison")

    r2_headers = ["Predictor", "wOBA Weight", "xwOBA Weight",
                   "R-Squared (Formula)", "R-Squared (Computed)", "N"]
    for c, h in enumerate(r2_headers, 1):
        ws_r2.cell(row=1, column=c, value=h)
    style_header_row(ws_r2, len(r2_headers))

    # Column references in Scatter Data
    prior_woba_col = "D"   # Prior wOBA
    prior_xwoba_col = "E"  # Prior xwOBA
    current_woba_col = "F"  # Current wOBA

    for r_idx, (_, row) in enumerate(comparison_df.iterrows(), 2):
        ws_r2.cell(row=r_idx, column=1, value=row["Predictor"])
        ws_r2.cell(row=r_idx, column=2, value=row["wOBA Weight"])
        ws_r2.cell(row=r_idx, column=2).number_format = "0.00"
        ws_r2.cell(row=r_idx, column=3, value=row["xwOBA Weight"])
        ws_r2.cell(row=r_idx, column=3).number_format = "0.00"

        # Build predictor formula reference
        w_woba = row["wOBA Weight"]
        w_xwoba = row["xwOBA Weight"]
        if w_xwoba == 0:
            pred_expr = f"'Scatter Data'!{prior_woba_col}2:{prior_woba_col}{last_scatter}"
        elif w_woba == 0:
            pred_expr = f"'Scatter Data'!{prior_xwoba_col}2:{prior_xwoba_col}{last_scatter}"
        else:
            pred_expr = (
                f"({w_woba}*'Scatter Data'!{prior_woba_col}2:{prior_woba_col}{last_scatter}"
                f"+{w_xwoba}*'Scatter Data'!{prior_xwoba_col}2:{prior_xwoba_col}{last_scatter})"
            )

        actual_range = f"'Scatter Data'!{current_woba_col}2:{current_woba_col}{last_scatter}"

        # R-squared via formula: 1 - SS_res/SS_tot
        r2_formula = (
            f"=1-SUMPRODUCT(({actual_range}-{pred_expr})^2)"
            f"/SUMPRODUCT(({actual_range}-AVERAGE({actual_range}))^2)"
        )
        ws_r2.cell(row=r_idx, column=4, value=r2_formula).number_format = "0.0000"

        ws_r2.cell(row=r_idx, column=5, value=round(row["R_squared"], 4))
        ws_r2.cell(row=r_idx, column=5).number_format = "0.0000"
        ws_r2.cell(row=r_idx, column=6, value=int(row["N"]))

    # Conditional formatting on R-squared (higher is better)
    n_blends = len(BLEND_RATIOS)
    rng_r2 = f"D2:D{1 + n_blends}"
    add_conditional_min_max(ws_r2, rng_r2, green_best="max")

    ws_r2.cell(row=1, column=4).comment = Comment(
        "R-squared measures how much variance in current-season wOBA\n"
        "is explained by the prior-season predictor.\n"
        "Formula: 1 - SS_residual / SS_total\n"
        "Higher R-squared = better predictor.",
        "Analysis Script",
    )

    auto_width(ws_r2)

    # ── Sheet 3: RMSE Comparison ──
    ws_rmse = wb.create_sheet("RMSE Comparison")

    rmse_headers = ["Predictor", "wOBA Weight", "xwOBA Weight",
                     "RMSE (Formula)", "RMSE (Computed)", "N"]
    for c, h in enumerate(rmse_headers, 1):
        ws_rmse.cell(row=1, column=c, value=h)
    style_header_row(ws_rmse, len(rmse_headers))

    for r_idx, (_, row) in enumerate(comparison_df.iterrows(), 2):
        ws_rmse.cell(row=r_idx, column=1, value=row["Predictor"])
        ws_rmse.cell(row=r_idx, column=2, value=row["wOBA Weight"])
        ws_rmse.cell(row=r_idx, column=2).number_format = "0.00"
        ws_rmse.cell(row=r_idx, column=3, value=row["xwOBA Weight"])
        ws_rmse.cell(row=r_idx, column=3).number_format = "0.00"

        w_woba = row["wOBA Weight"]
        w_xwoba = row["xwOBA Weight"]
        if w_xwoba == 0:
            pred_expr = f"'Scatter Data'!{prior_woba_col}2:{prior_woba_col}{last_scatter}"
        elif w_woba == 0:
            pred_expr = f"'Scatter Data'!{prior_xwoba_col}2:{prior_xwoba_col}{last_scatter}"
        else:
            pred_expr = (
                f"({w_woba}*'Scatter Data'!{prior_woba_col}2:{prior_woba_col}{last_scatter}"
                f"+{w_xwoba}*'Scatter Data'!{prior_xwoba_col}2:{prior_xwoba_col}{last_scatter})"
            )
        actual_range = f"'Scatter Data'!{current_woba_col}2:{current_woba_col}{last_scatter}"

        rmse_formula = (
            f"=SQRT(SUMPRODUCT(({actual_range}-{pred_expr})^2)"
            f"/COUNTA({actual_range}))"
        )
        ws_rmse.cell(row=r_idx, column=4, value=rmse_formula).number_format = "0.000000"
        ws_rmse.cell(row=r_idx, column=5, value=round(row["RMSE"], 6))
        ws_rmse.cell(row=r_idx, column=5).number_format = "0.000000"
        ws_rmse.cell(row=r_idx, column=6, value=int(row["N"]))

    rng_rmse = f"D2:D{1 + n_blends}"
    add_conditional_min_max(ws_rmse, rng_rmse, green_best="min")

    auto_width(ws_rmse)

    # ── Sheet 4: Optimal Blend ──
    ws_opt = wb.create_sheet("Optimal Blend")
    ws_opt.cell(row=1, column=1, value="Optimal xwOBA/wOBA Blend").font = Font(
        bold=True, size=14
    )

    ws_opt.cell(row=3, column=1, value="Best Predictor:").font = HEADER_FONT
    ws_opt.cell(
        row=3, column=2,
        value=f"=INDEX('RMSE Comparison'!A2:A{1+n_blends},"
        f"MATCH(MIN('RMSE Comparison'!D2:D{1+n_blends}),"
        f"'RMSE Comparison'!D2:D{1+n_blends},0))",
    )
    ws_opt.cell(row=3, column=2).comment = Comment(
        "The blend ratio with the lowest RMSE for predicting\n"
        "next-season wOBA from prior-season data.",
        "Analysis Script",
    )

    ws_opt.cell(row=4, column=1, value="Minimum RMSE:").font = HEADER_FONT
    ws_opt.cell(
        row=4, column=2,
        value=f"=MIN('RMSE Comparison'!D2:D{1+n_blends})",
    )
    ws_opt.cell(row=4, column=2).number_format = "0.000000"

    ws_opt.cell(row=5, column=1, value="Best R-Squared:").font = HEADER_FONT
    ws_opt.cell(
        row=5, column=2,
        value=f"=MAX('R-Squared Comparison'!D2:D{1+n_blends})",
    )
    ws_opt.cell(row=5, column=2).number_format = "0.0000"

    ws_opt.cell(row=7, column=1, value="Key Insight:").font = HEADER_FONT
    ws_opt.cell(row=8, column=1, value=(
        "If xwOBA-heavy blends outperform wOBA-only, it confirms that Statcast\n"
        "expected metrics strip out luck/variance and better capture true talent.\n"
        "The optimal blend ratio should be used for in-season projection adjustments."
    )).alignment = Alignment(wrap_text=True)

    ws_opt.column_dimensions["A"].width = 25
    ws_opt.column_dimensions["B"].width = 30

    # ── Sheet 5: By Season ──
    ws_season = wb.create_sheet("By Season")

    if not season_df.empty:
        # Pivot: rows=season, columns=predictor, values=RMSE
        season_headers = ["Season"] + [label for _, _, label in BLEND_RATIOS]
        for c, h in enumerate(season_headers, 1):
            ws_season.cell(row=1, column=c, value=f"RMSE: {h}" if c > 1 else h)
        style_header_row(ws_season, len(season_headers))

        seasons_list = sorted(season_df["Season"].unique())
        for r_idx, season in enumerate(seasons_list, 2):
            ws_season.cell(row=r_idx, column=1, value=int(season))
            for c_idx, (_, _, label) in enumerate(BLEND_RATIOS, 2):
                match = season_df[
                    (season_df["Season"] == season) & (season_df["Predictor"] == label)
                ]
                if not match.empty:
                    ws_season.cell(
                        row=r_idx, column=c_idx, value=round(match.iloc[0]["RMSE"], 6)
                    ).number_format = "0.000000"

        # Conditional formatting per row (lower RMSE is better)
        for r_idx in range(2, 2 + len(seasons_list)):
            rng = f"B{r_idx}:{get_column_letter(1 + len(BLEND_RATIOS))}{r_idx}"
            add_conditional_min_max(ws_season, rng, green_best="min")

        # Add R-squared section below
        offset = 2 + len(seasons_list) + 1
        ws_season.cell(row=offset, column=1, value="R-Squared by Season").font = Font(
            bold=True, size=13
        )
        for c, h in enumerate(season_headers, 1):
            ws_season.cell(row=offset + 1, column=c, value=f"R2: {h}" if c > 1 else h)
        style_header_row(ws_season, len(season_headers), row=offset + 1)

        for r_idx, season in enumerate(seasons_list, offset + 2):
            ws_season.cell(row=r_idx, column=1, value=int(season))
            for c_idx, (_, _, label) in enumerate(BLEND_RATIOS, 2):
                match = season_df[
                    (season_df["Season"] == season) & (season_df["Predictor"] == label)
                ]
                if not match.empty:
                    ws_season.cell(
                        row=r_idx, column=c_idx, value=round(match.iloc[0]["R_squared"], 4)
                    ).number_format = "0.0000"
    else:
        ws_season.cell(row=1, column=1, value="No season-level data available.")

    auto_width(ws_season)

    wb.save(output_path)
    logger.info("Saved workbook to %s", output_path)


def main() -> None:
    """Entry point for xwOBA regression analysis."""
    parser = argparse.ArgumentParser(
        description="Analyze xwOBA vs wOBA as future performance predictor"
    )
    parser.add_argument(
        "--db-path",
        default=str(PROJECT_ROOT / "backtest_data.sqlite"),
        help="Path to backtest SQLite database",
    )
    parser.add_argument(
        "--output-dir",
        default=str(PROJECT_ROOT / "analysis"),
        help="Directory for Excel output",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(name)s %(levelname)s %(message)s",
    )

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    batting, statcast, player_ids = load_data(args.db_path)

    if batting.empty or statcast.empty:
        logger.warning("Insufficient data. Generating empty workbook.")
        wb = Workbook()
        wb.active.title = "No Data"
        wb.active.cell(row=1, column=1, value="Insufficient data in database.")
        wb.save(str(output_dir / "xwoba_regression_analysis.xlsx"))
        return

    results = run_analysis(batting, statcast, player_ids)
    write_excel(results, str(output_dir / "xwoba_regression_analysis.xlsx"))
    logger.info("xwOBA regression analysis complete.")


if __name__ == "__main__":
    main()
