"""Shared utilities for analysis scripts.

Provides common Excel styling, formatting helpers, and database loading
used across all analysis scripts.
"""

import logging
import re
import sqlite3
import unicodedata
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent

# ── Excel Style Constants ──

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def auto_width(ws: Any, max_col_width: int = 40) -> None:
    """Auto-size columns based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_col_width)


def style_header_row(ws: Any, num_cols: int, row: int = 1) -> None:
    """Apply header styling to a row."""
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def add_conditional_min_max(
    ws: Any,
    cell_range: str,
    *,
    green_best: str = "min",
) -> None:
    """Add conditional formatting: green=best, red=worst.

    Args:
        ws: Worksheet to format.
        cell_range: e.g. "B2:B8".
        green_best: "min" if lower is better (RMSE), "max" if higher is better (R-squared).
    """
    if green_best == "min":
        green_formula = f"MIN({cell_range})"
        red_formula = f"MAX({cell_range})"
    else:
        green_formula = f"MAX({cell_range})"
        red_formula = f"MIN({cell_range})"

    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=[green_formula], fill=GREEN_FILL),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="equal", formula=[red_formula], fill=RED_FILL),
    )


def get_db_connection(db_path: str) -> sqlite3.Connection:
    """Open a SQLite connection, raising FileNotFoundError if the DB is missing."""
    path = Path(db_path)
    if not path.exists():
        raise FileNotFoundError(
            f"Database not found: {db_path}\n"
            "Run 'uv run python -m scripts.data_pipeline' first to create "
            "the backtest database."
        )
    return sqlite3.connect(db_path)


# ── Player ID Crosswalk ──


def _normalize_name(name: str) -> str:
    """Normalize a player name to lowercase 'first last' for matching.

    Handles FanGraphs format ("First Last") and Statcast format ("Last, First").
    Strips accents, suffixes (Jr., Sr., II, III, IV), and extra whitespace.
    """
    if not isinstance(name, str) or not name.strip():
        return ""
    # Strip accent characters
    name = unicodedata.normalize("NFD", name)
    name = "".join(c for c in name if unicodedata.category(c) != "Mn")
    # Remove common suffixes
    name = re.sub(r"\b(jr\.?|sr\.?|ii|iii|iv)\b", "", name, flags=re.IGNORECASE)
    name = name.strip().strip(",").strip()
    # Handle "Last, First" format
    if "," in name:
        parts = [p.strip() for p in name.split(",", 1)]
        if len(parts) == 2 and parts[0] and parts[1]:
            name = f"{parts[1]} {parts[0]}"
    return " ".join(name.lower().split())


def load_player_id_crosswalk(
    db_path: str,
    *,
    columns: list[str] | None = None,
) -> pd.DataFrame:
    """Load fangraphs_id <-> mlbam_id crosswalk with multi-source fallback.

    Tries in order:
    1. ``player_ids`` table in the backtest database (Chadwick crosswalk).
    2. ``players`` table in the live app database (fantasy_baseball.db).
    3. Name-based join between ``batting_season`` and ``statcast_season``
       in the backtest database.

    Args:
        db_path: Path to the backtest SQLite database.
        columns: Extra columns to include beyond fangraphs_id / mlbam_id.
            Currently only ``["name_first", "name_last"]`` is supported for
            source 1 (Chadwick); other sources return only the two ID columns.

    Returns:
        DataFrame with at least ``fangraphs_id`` and ``mlbam_id`` columns
        (both as str). May be empty if all sources fail.
    """
    # --- Source 1: player_ids table in backtest DB ---
    conn = get_db_connection(db_path)
    try:
        extra = ""
        if columns:
            safe = [c for c in columns if c not in ("fangraphs_id", "mlbam_id")]
            if safe:
                extra = ", " + ", ".join(safe)
        player_ids = pd.read_sql_query(
            f"""
            SELECT fangraphs_id, mlbam_id{extra}
            FROM player_ids
            WHERE fangraphs_id IS NOT NULL AND mlbam_id IS NOT NULL
            """,
            conn,
        )
    except Exception:
        player_ids = pd.DataFrame()
    conn.close()

    if not player_ids.empty:
        player_ids["fangraphs_id"] = player_ids["fangraphs_id"].astype(str)
        player_ids["mlbam_id"] = player_ids["mlbam_id"].astype(str)
        logger.info(
            "Loaded %d ID mappings from player_ids table (backtest DB)",
            len(player_ids),
        )
        return player_ids

    # --- Source 2: players table in the live app DB ---
    app_db = PROJECT_ROOT / "fantasy_baseball.db"
    if app_db.exists():
        try:
            app_conn = sqlite3.connect(str(app_db))
            player_ids = pd.read_sql_query(
                """
                SELECT fangraphs_id, mlbam_id
                FROM players
                WHERE fangraphs_id IS NOT NULL AND mlbam_id IS NOT NULL
                """,
                app_conn,
            )
            app_conn.close()
        except Exception:
            player_ids = pd.DataFrame()

        if not player_ids.empty:
            player_ids["fangraphs_id"] = player_ids["fangraphs_id"].astype(str)
            player_ids["mlbam_id"] = player_ids["mlbam_id"].astype(str)
            logger.info(
                "Loaded %d ID mappings from fantasy_baseball.db players table",
                len(player_ids),
            )
            return player_ids

    # --- Source 3: Name-based join ---
    logger.info(
        "player_ids crosswalk empty and app DB unavailable; "
        "falling back to name-based join"
    )
    conn = get_db_connection(db_path)
    try:
        batting_names = pd.read_sql_query(
            """
            SELECT DISTINCT fangraphs_id, name
            FROM batting_season
            WHERE fangraphs_id IS NOT NULL AND name IS NOT NULL
            """,
            conn,
        )
        statcast_names = pd.read_sql_query(
            """
            SELECT DISTINCT mlbam_id, name
            FROM statcast_season
            WHERE mlbam_id IS NOT NULL AND name IS NOT NULL
            """,
            conn,
        )
    except Exception:
        batting_names = pd.DataFrame()
        statcast_names = pd.DataFrame()
    conn.close()

    if batting_names.empty or statcast_names.empty:
        logger.warning("Name-based fallback failed: no batting or statcast names found")
        return pd.DataFrame(columns=["fangraphs_id", "mlbam_id"])

    batting_names["norm_name"] = batting_names["name"].apply(_normalize_name)
    statcast_names["norm_name"] = statcast_names["name"].apply(_normalize_name)

    # Drop empty / duplicate normalized names
    batting_names = batting_names[batting_names["norm_name"] != ""].drop_duplicates(
        subset=["norm_name"]
    )
    statcast_names = statcast_names[statcast_names["norm_name"] != ""].drop_duplicates(
        subset=["norm_name"]
    )

    merged = batting_names.merge(statcast_names, on="norm_name", how="inner")
    player_ids = merged[["fangraphs_id", "mlbam_id"]].copy()
    player_ids["fangraphs_id"] = player_ids["fangraphs_id"].astype(str)
    player_ids["mlbam_id"] = player_ids["mlbam_id"].astype(str)

    logger.info(
        "Name-based fallback produced %d ID mappings", len(player_ids)
    )
    return player_ids
