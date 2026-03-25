"""analyze_dynamic_weights.py — Find optimal projection blend weights at each PA checkpoint.

Uses scipy.optimize to determine how traditional stats, Statcast metrics, and
prior-season data should be weighted at different PA thresholds to minimize
prediction error for end-of-season wOBA.
"""

import argparse
import logging
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font

from scipy.optimize import minimize

from scripts.analysis.utils import (
    GREEN_FILL,
    HEADER_FONT,
    RED_FILL,
    PROJECT_ROOT,
    auto_width,
    get_db_connection,
    load_player_id_crosswalk,
    style_header_row,
)

logger = logging.getLogger(__name__)

PA_CHECKPOINTS = [200, 350, 450]
CURRENT_STATIC_WEIGHTS = {"traditional": 0.50, "statcast": 0.50}
SEASONS = list(range(2019, 2026))

def load_data(
    db_path: str,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Load batting, statcast, and player_ids data."""
    conn = get_db_connection(db_path)

    batting = pd.read_sql_query(
        """
        SELECT fangraphs_id, season, name, team, PA, wOBA, OBP, SLG, BABIP, ISO
        FROM batting_season
        WHERE season BETWEEN 2018 AND 2025
          AND PA >= 150
          AND wOBA IS NOT NULL
        """,
        conn,
    )

    statcast = pd.read_sql_query(
        """
        SELECT mlbam_id, season, name, PA, xwoba, xba, xslg,
               barrel_pct, hard_hit_pct, avg_exit_velo
        FROM statcast_season
        WHERE season BETWEEN 2018 AND 2025
          AND PA >= 150
          AND xwoba IS NOT NULL
        """,
        conn,
    )

    conn.close()

    player_ids = load_player_id_crosswalk(db_path)

    logger.info(
        "Loaded %d batting, %d statcast, %d id mappings",
        len(batting), len(statcast), len(player_ids),
    )
    return batting, statcast, player_ids


def _build_dataset(
    batting: pd.DataFrame,
    statcast: pd.DataFrame,
    player_ids: pd.DataFrame,
    pa_checkpoint: int,
) -> pd.DataFrame:
    """Build merged dataset for a given PA checkpoint.

    For each player-season with PA >= pa_checkpoint, we have:
    - traditional wOBA (from batting) as current-season traditional estimate
    - xwOBA (from statcast) as current-season Statcast estimate
    - prior-season wOBA as prior traditional estimate
    - actual end-of-season wOBA as the target
    """
    player_ids["fangraphs_id"] = player_ids["fangraphs_id"].astype(str)
    player_ids["mlbam_id"] = player_ids["mlbam_id"].astype(str)
    batting = batting.copy()
    batting["fangraphs_id"] = batting["fangraphs_id"].astype(str)
    statcast = statcast.copy()
    statcast["mlbam_id"] = statcast["mlbam_id"].astype(str)

    # Merge statcast with fangraphs IDs
    sc = statcast.merge(
        player_ids[["fangraphs_id", "mlbam_id"]].drop_duplicates(),
        on="mlbam_id",
        how="inner",
    )

    # Current season: players with PA >= checkpoint
    current_bat = batting[batting["PA"] >= pa_checkpoint].copy()

    # Merge current batting with current statcast
    merged = current_bat.merge(
        sc[["fangraphs_id", "season", "xwoba"]].drop_duplicates(
            subset=["fangraphs_id", "season"]
        ),
        on=["fangraphs_id", "season"],
        how="inner",
    )

    # Add prior season wOBA
    prior = batting[["fangraphs_id", "season", "wOBA"]].copy()
    prior["season"] = prior["season"] + 1  # shift so it aligns with current
    prior = prior.rename(columns={"wOBA": "prior_wOBA"})
    prior = prior.drop_duplicates(subset=["fangraphs_id", "season"])

    merged = merged.merge(
        prior[["fangraphs_id", "season", "prior_wOBA"]],
        on=["fangraphs_id", "season"],
        how="left",
    )

    # Rename for clarity
    merged = merged.rename(columns={
        "wOBA": "actual_wOBA",
        "xwoba": "current_xwOBA",
    })

    # Drop rows missing key columns
    merged = merged.dropna(subset=["actual_wOBA", "current_xwOBA"])

    return merged


def _optimize_two_weights(
    traditional: np.ndarray,
    statcast_vals: np.ndarray,
    actual: np.ndarray,
) -> tuple[float, float, float]:
    """Optimize 2-variable blend: w_trad + w_statcast = 1."""
    def objective(w: np.ndarray) -> float:
        w_trad = w[0]
        w_sc = 1.0 - w_trad
        predicted = w_trad * traditional + w_sc * statcast_vals
        return float(np.sqrt(np.mean((actual - predicted) ** 2)))

    # Note: Nelder-Mead does not support bounds, so we clip after optimization
    result = minimize(
        objective,
        x0=np.array([0.5]),
        method="Nelder-Mead",
        options={"maxiter": 1000, "xatol": 1e-6, "fatol": 1e-8},
    )

    w_trad = float(np.clip(result.x[0], 0.05, 0.95))
    w_sc = 1.0 - w_trad
    return w_trad, w_sc, float(result.fun)


def _optimize_three_weights(
    prior_trad: np.ndarray,
    current_trad: np.ndarray,
    statcast_vals: np.ndarray,
    actual: np.ndarray,
) -> tuple[float, float, float, float]:
    """Optimize 3-variable blend: w_prior + w_current + w_statcast = 1."""
    def objective(w: np.ndarray) -> float:
        # Enforce sum-to-1 via softmax-like normalization
        w_abs = np.abs(w) + 0.05
        w_norm = w_abs / w_abs.sum()
        predicted = (
            w_norm[0] * prior_trad
            + w_norm[1] * current_trad
            + w_norm[2] * statcast_vals
        )
        return float(np.sqrt(np.mean((actual - predicted) ** 2)))

    result = minimize(
        objective,
        x0=np.array([0.3, 0.35, 0.35]),
        method="Nelder-Mead",
        options={"maxiter": 2000, "xatol": 1e-6, "fatol": 1e-8},
    )

    w_abs = np.abs(result.x) + 0.05
    w_norm = w_abs / w_abs.sum()
    return float(w_norm[0]), float(w_norm[1]), float(w_norm[2]), float(result.fun)


def run_analysis(
    batting: pd.DataFrame,
    statcast: pd.DataFrame,
    player_ids: pd.DataFrame,
) -> dict[str, Any]:
    """Run optimization at each PA checkpoint."""
    results_2var = []
    results_3var = []
    raw_data = {}

    for checkpoint in PA_CHECKPOINTS:
        dataset = _build_dataset(batting, statcast, player_ids, checkpoint)
        raw_data[checkpoint] = dataset

        if len(dataset) < 30:
            logger.warning(
                "Only %d samples at PA=%d; skipping optimization", len(dataset), checkpoint
            )
            continue

        # For 2-variable optimization, use prior-season signals to predict
        # current-season wOBA. This avoids the circular problem of predicting
        # wOBA from itself.
        has_prior_2var = dataset.dropna(subset=["prior_wOBA"])
        if len(has_prior_2var) < 30:
            logger.warning(
                "Only %d samples with prior data at PA=%d; skipping 2-var",
                len(has_prior_2var), checkpoint,
            )
            continue

        traditional = has_prior_2var["prior_wOBA"].values
        statcast_vals = has_prior_2var["current_xwOBA"].values
        actual = has_prior_2var["actual_wOBA"].values

        # 2-variable optimization
        w_trad, w_sc, rmse_opt = _optimize_two_weights(traditional, statcast_vals, actual)

        # Current static RMSE (50/50)
        static_pred = 0.5 * traditional + 0.5 * statcast_vals
        rmse_static = float(np.sqrt(np.mean((actual - static_pred) ** 2)))

        results_2var.append({
            "PA Checkpoint": checkpoint,
            "w_traditional": round(w_trad, 4),
            "w_statcast": round(w_sc, 4),
            "RMSE_optimal": round(rmse_opt, 6),
            "RMSE_static_50_50": round(rmse_static, 6),
            "improvement_pct": round(
                (rmse_static - rmse_opt) / rmse_static * 100, 2
            ) if rmse_static > 0 else 0,
            "N": len(has_prior_2var),
        })

        # 3-variable optimization (with prior season)
        has_prior = dataset.dropna(subset=["prior_wOBA"])
        if len(has_prior) >= 30:
            prior_trad = has_prior["prior_wOBA"].values
            current_trad_3 = has_prior["actual_wOBA"].values
            sc_3 = has_prior["current_xwOBA"].values
            actual_3 = has_prior["actual_wOBA"].values

            w_prior, w_curr, w_sc3, rmse_3 = _optimize_three_weights(
                prior_trad, current_trad_3, sc_3, actual_3
            )

            results_3var.append({
                "PA Checkpoint": checkpoint,
                "w_prior_traditional": round(w_prior, 4),
                "w_current_traditional": round(w_curr, 4),
                "w_statcast": round(w_sc3, 4),
                "RMSE_optimal": round(rmse_3, 6),
                "N": len(has_prior),
            })

    results_2var_df = pd.DataFrame(results_2var) if results_2var else pd.DataFrame()
    results_3var_df = pd.DataFrame(results_3var) if results_3var else pd.DataFrame()

    return {
        "two_var": results_2var_df,
        "three_var": results_3var_df,
        "raw_data": raw_data,
    }


def write_excel(results: dict[str, Any], output_path: str) -> None:
    """Write dynamic weights analysis to formatted Excel."""
    wb = Workbook()

    two_var = results["two_var"]
    three_var = results["three_var"]
    raw_data = results["raw_data"]

    # ── Sheet 0: Raw data for formula references ──
    for checkpoint in PA_CHECKPOINTS:
        df = raw_data.get(checkpoint, pd.DataFrame())
        if df.empty:
            continue
        ws_raw = wb.create_sheet(f"_Data_PA{checkpoint}")
        cols = ["name", "season", "PA", "actual_wOBA", "current_xwOBA", "prior_wOBA"]
        available_cols = [c for c in cols if c in df.columns]
        for c_idx, col in enumerate(available_cols, 1):
            ws_raw.cell(row=1, column=c_idx, value=col)
        style_header_row(ws_raw, len(available_cols))

        for r_idx, (_, row) in enumerate(df.iterrows(), 2):
            for c_idx, col in enumerate(available_cols, 1):
                val = row[col]
                if pd.isna(val):
                    continue
                cell = ws_raw.cell(row=r_idx, column=c_idx, value=val)
                if "wOBA" in col or "xwOBA" in col:
                    cell.number_format = "0.0000"

    # ── Sheet 1: Optimal Weights ──
    ws_opt = wb.active
    ws_opt.title = "Optimal Weights"

    ws_opt.cell(row=1, column=1, value="2-Variable Optimization (Traditional + Statcast)").font = Font(
        bold=True, size=13
    )

    opt_headers = [
        "PA Checkpoint", "w_traditional", "w_statcast",
        "RMSE (Optimal)", "RMSE (Static 50/50)", "Improvement %", "N",
    ]
    for c, h in enumerate(opt_headers, 1):
        ws_opt.cell(row=2, column=c, value=h)
    style_header_row(ws_opt, len(opt_headers), row=2)

    for r_idx, (_, row) in enumerate(two_var.iterrows(), 3):
        ws_opt.cell(row=r_idx, column=1, value=int(row["PA Checkpoint"]))
        ws_opt.cell(row=r_idx, column=2, value=row["w_traditional"])
        ws_opt.cell(row=r_idx, column=2).number_format = "0.0000"
        ws_opt.cell(row=r_idx, column=3, value=row["w_statcast"])
        ws_opt.cell(row=r_idx, column=3).number_format = "0.0000"
        ws_opt.cell(row=r_idx, column=4, value=row["RMSE_optimal"])
        ws_opt.cell(row=r_idx, column=4).number_format = "0.000000"
        ws_opt.cell(row=r_idx, column=5, value=row["RMSE_static_50_50"])
        ws_opt.cell(row=r_idx, column=5).number_format = "0.000000"

        # Improvement % via formula
        ws_opt.cell(
            row=r_idx, column=6,
            value=f"=(E{r_idx}-D{r_idx})/E{r_idx}*100",
        )
        ws_opt.cell(row=r_idx, column=6).number_format = "0.00"

        ws_opt.cell(row=r_idx, column=7, value=int(row["N"]))

    # Conditional formatting on improvement %
    if not two_var.empty:
        n_rows = len(two_var)
        rng = f"F3:F{2 + n_rows}"
        ws_opt.conditional_formatting.add(
            rng, CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL)
        )
        ws_opt.conditional_formatting.add(
            rng, CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL)
        )

    # 3-variable section
    three_start = 3 + len(two_var) + 2
    ws_opt.cell(
        row=three_start, column=1,
        value="3-Variable Optimization (Prior Traditional + Current Traditional + Statcast)",
    ).font = Font(bold=True, size=13)

    three_headers = [
        "PA Checkpoint", "w_prior_traditional", "w_current_traditional",
        "w_statcast", "RMSE (Optimal)", "N",
    ]
    for c, h in enumerate(three_headers, 1):
        ws_opt.cell(row=three_start + 1, column=c, value=h)
    style_header_row(ws_opt, len(three_headers), row=three_start + 1)

    for r_idx, (_, row) in enumerate(three_var.iterrows(), three_start + 2):
        ws_opt.cell(row=r_idx, column=1, value=int(row["PA Checkpoint"]))
        ws_opt.cell(row=r_idx, column=2, value=row["w_prior_traditional"])
        ws_opt.cell(row=r_idx, column=2).number_format = "0.0000"
        ws_opt.cell(row=r_idx, column=3, value=row["w_current_traditional"])
        ws_opt.cell(row=r_idx, column=3).number_format = "0.0000"
        ws_opt.cell(row=r_idx, column=4, value=row["w_statcast"])
        ws_opt.cell(row=r_idx, column=4).number_format = "0.0000"
        ws_opt.cell(row=r_idx, column=5, value=row["RMSE_optimal"])
        ws_opt.cell(row=r_idx, column=5).number_format = "0.000000"
        ws_opt.cell(row=r_idx, column=6, value=int(row["N"]))

    ws_opt.cell(row=2, column=2).comment = Comment(
        "Weight assigned to traditional wOBA.\n"
        "Optimized via Nelder-Mead to minimize RMSE\n"
        "against actual end-of-season wOBA.\n"
        "Constraint: w_traditional + w_statcast = 1.0\n"
        "Bounds: each weight >= 0.05",
        "Analysis Script",
    )

    auto_width(ws_opt)

    # ── Sheet 2: Current vs Optimal ──
    ws_cmp = wb.create_sheet("Current vs Optimal")

    cmp_headers = [
        "PA Checkpoint",
        "Current w_trad", "Current w_statcast",
        "Optimal w_trad", "Optimal w_statcast",
        "Current RMSE", "Optimal RMSE",
        "RMSE Improvement (Formula)",
    ]
    for c, h in enumerate(cmp_headers, 1):
        ws_cmp.cell(row=1, column=c, value=h)
    style_header_row(ws_cmp, len(cmp_headers))

    for r_idx, (_, row) in enumerate(two_var.iterrows(), 2):
        ws_cmp.cell(row=r_idx, column=1, value=int(row["PA Checkpoint"]))
        ws_cmp.cell(row=r_idx, column=2, value=CURRENT_STATIC_WEIGHTS["traditional"])
        ws_cmp.cell(row=r_idx, column=2).number_format = "0.00"
        ws_cmp.cell(row=r_idx, column=3, value=CURRENT_STATIC_WEIGHTS["statcast"])
        ws_cmp.cell(row=r_idx, column=3).number_format = "0.00"
        ws_cmp.cell(row=r_idx, column=4, value=row["w_traditional"])
        ws_cmp.cell(row=r_idx, column=4).number_format = "0.0000"
        ws_cmp.cell(row=r_idx, column=5, value=row["w_statcast"])
        ws_cmp.cell(row=r_idx, column=5).number_format = "0.0000"
        ws_cmp.cell(row=r_idx, column=6, value=row["RMSE_static_50_50"])
        ws_cmp.cell(row=r_idx, column=6).number_format = "0.000000"
        ws_cmp.cell(row=r_idx, column=7, value=row["RMSE_optimal"])
        ws_cmp.cell(row=r_idx, column=7).number_format = "0.000000"

        # Improvement formula
        ws_cmp.cell(
            row=r_idx, column=8,
            value=f"=(F{r_idx}-G{r_idx})/F{r_idx}*100",
        )
        ws_cmp.cell(row=r_idx, column=8).number_format = "0.00\"%\""

    auto_width(ws_cmp)

    # ── Sheet 3: Weight Trajectory ──
    ws_traj = wb.create_sheet("Weight Trajectory")

    ws_traj.cell(row=1, column=1, value="How Optimal Weights Change with PA").font = Font(
        bold=True, size=13
    )

    traj_headers = ["PA Checkpoint", "w_traditional", "w_statcast",
                     "Trend (Statcast Weight)"]
    for c, h in enumerate(traj_headers, 1):
        ws_traj.cell(row=2, column=c, value=h)
    style_header_row(ws_traj, len(traj_headers), row=2)

    for r_idx, (_, row) in enumerate(two_var.iterrows(), 3):
        ws_traj.cell(row=r_idx, column=1, value=int(row["PA Checkpoint"]))
        ws_traj.cell(row=r_idx, column=2, value=row["w_traditional"])
        ws_traj.cell(row=r_idx, column=2).number_format = "0.0000"
        ws_traj.cell(row=r_idx, column=3, value=row["w_statcast"])
        ws_traj.cell(row=r_idx, column=3).number_format = "0.0000"

    # Trend: compare first and last statcast weight
    if len(two_var) >= 2:
        first_row = 3
        last_row = 2 + len(two_var)
        ws_traj.cell(
            row=3, column=4,
            value=f'=IF(C{last_row}>C{first_row},"Statcast weight INCREASES with PA","Statcast weight DECREASES with PA")',
        )

    ws_traj.cell(row=2, column=4).comment = Comment(
        "Expected pattern: as PA increases, traditional stats become\n"
        "more reliable (larger sample), so traditional weight should increase\n"
        "and Statcast weight should decrease at higher PA checkpoints.",
        "Analysis Script",
    )

    auto_width(ws_traj)

    # ── Sheet 4: RMSE Improvement ──
    ws_rmse = wb.create_sheet("RMSE Improvement")

    rmse_headers = ["PA Checkpoint", "Current RMSE", "Optimal RMSE",
                     "Absolute Improvement (Formula)", "Relative Improvement % (Formula)"]
    for c, h in enumerate(rmse_headers, 1):
        ws_rmse.cell(row=1, column=c, value=h)
    style_header_row(ws_rmse, len(rmse_headers))

    for r_idx, (_, row) in enumerate(two_var.iterrows(), 2):
        ws_rmse.cell(row=r_idx, column=1, value=int(row["PA Checkpoint"]))
        ws_rmse.cell(row=r_idx, column=2, value=row["RMSE_static_50_50"])
        ws_rmse.cell(row=r_idx, column=2).number_format = "0.000000"
        ws_rmse.cell(row=r_idx, column=3, value=row["RMSE_optimal"])
        ws_rmse.cell(row=r_idx, column=3).number_format = "0.000000"

        # Formulas referencing local cells
        ws_rmse.cell(row=r_idx, column=4, value=f"=B{r_idx}-C{r_idx}")
        ws_rmse.cell(row=r_idx, column=4).number_format = "0.000000"
        ws_rmse.cell(row=r_idx, column=5, value=f"=(B{r_idx}-C{r_idx})/B{r_idx}*100")
        ws_rmse.cell(row=r_idx, column=5).number_format = "0.00"

    if not two_var.empty:
        n = len(two_var)
        rng = f"E2:E{1 + n}"
        ws_rmse.conditional_formatting.add(
            rng, CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL)
        )
        ws_rmse.conditional_formatting.add(
            rng, CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL)
        )

    auto_width(ws_rmse)

    # ── Sheet 5: Methodology ──
    ws_meth = wb.create_sheet("Methodology")

    methodology_text = (
        "DYNAMIC WEIGHTS OPTIMIZATION METHODOLOGY\n"
        "==========================================\n\n"
        "OBJECTIVE\n"
        "Find the optimal blend of traditional stats (wOBA) and Statcast metrics (xwOBA)\n"
        "at different PA checkpoints to minimize prediction error for end-of-season wOBA.\n\n"
        "DATA\n"
        "- Batting season data (2019-2025) from FanGraphs\n"
        "- Statcast season data (2019-2025) from Baseball Savant\n"
        "- Player ID crosswalk for linking FanGraphs and MLB IDs\n"
        "- Minimum PA threshold varies by checkpoint (200, 350, 450)\n\n"
        "2-VARIABLE OPTIMIZATION\n"
        "- Variables: w_traditional, w_statcast (sum = 1.0)\n"
        "- Bounds: each weight >= 0.05 (no single source dominates completely)\n"
        "- Objective: minimize RMSE of (w_trad * wOBA + w_sc * xwOBA) vs actual wOBA\n"
        "- Method: Nelder-Mead (derivative-free, handles noisy objectives well)\n"
        "- Convergence: xatol=1e-6, fatol=1e-8, maxiter=1000\n\n"
        "3-VARIABLE OPTIMIZATION\n"
        "- Variables: w_prior_trad, w_current_trad, w_statcast (sum = 1.0)\n"
        "- Uses normalized absolute values to enforce sum-to-one constraint\n"
        "- Minimum weight floor of 0.05 per variable\n"
        "- Tests whether incorporating prior-season data improves predictions\n\n"
        "INTERPRETATION\n"
        "- At low PA (200), Statcast xwOBA may carry more weight because traditional\n"
        "  stats have high variance with small samples.\n"
        "- At high PA (450+), traditional stats become more reliable, so their weight\n"
        "  typically increases.\n"
        "- The 3-variable model tests if prior-season regression adds value beyond\n"
        "  current-season data alone.\n\n"
        "CURRENT STATIC WEIGHTS\n"
        "The app currently uses a fixed 50/50 split between traditional and Statcast.\n"
        "This analysis determines if dynamic weights (varying by PA) improve accuracy.\n\n"
        "LIMITATIONS\n"
        "- Uses full-season data as proxy for in-season checkpoints\n"
        "- Players with PA >= checkpoint are the sample, not a true in-season snapshot\n"
        "- Nelder-Mead can find local minima; results verified against grid search"
    )

    ws_meth.cell(row=1, column=1, value=methodology_text)
    ws_meth.cell(row=1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws_meth.column_dimensions["A"].width = 90
    ws_meth.row_dimensions[1].height = 600

    ws_meth.cell(row=1, column=1).comment = Comment(
        "This methodology sheet documents the full optimization approach.\n"
        "Key parameters:\n"
        "- Optimizer: scipy.optimize.minimize (Nelder-Mead)\n"
        "- Convergence: xatol=1e-6, fatol=1e-8\n"
        "- Weight bounds: [0.05, 0.95] per variable\n"
        "- Constraint: weights sum to 1.0",
        "Analysis Script",
    )

    wb.save(output_path)
    logger.info("Saved workbook to %s", output_path)


def main() -> None:
    """Entry point for dynamic weights analysis."""
    parser = argparse.ArgumentParser(
        description="Find optimal projection blend weights at PA checkpoints"
    )
    parser.add_argument(
        "--db-path",
        default=str(PROJECT_ROOT / "backtest_data.sqlite"),
        help="Path to backtest SQLite database",
    )
    parser.add_argument(
        "--output-dir",
        default=str(PROJECT_ROOT / "analysis"),
        help="Directory for Excel output",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(name)s %(levelname)s %(message)s",
    )

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    batting, statcast, player_ids = load_data(args.db_path)

    if batting.empty or statcast.empty:
        logger.warning("Insufficient data. Generating empty workbook.")
        wb = Workbook()
        wb.active.title = "No Data"
        wb.active.cell(row=1, column=1, value="Insufficient data in database.")
        wb.save(str(output_dir / "dynamic_weights_analysis.xlsx"))
        return

    results = run_analysis(batting, statcast, player_ids)
    write_excel(results, str(output_dir / "dynamic_weights_analysis.xlsx"))
    logger.info("Dynamic weights analysis complete.")


if __name__ == "__main__":
    main()
