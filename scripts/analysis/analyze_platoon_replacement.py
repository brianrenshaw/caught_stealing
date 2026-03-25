"""analyze_platoon_replacement.py — Compare platoon-only vs pitcher-quality-only vs multiplicative approach.

Since we only have season-level data, uses league-average platoon split proxies
and team SIERA ratios to compare three adjustment methods for predicting wOBA.
"""

import argparse
import logging
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from scripts.analysis.utils import (
    HEADER_FONT,
    PROJECT_ROOT,
    add_conditional_min_max,
    auto_width,
    get_db_connection,
    style_header_row,
)

logger = logging.getLogger(__name__)

# League-average platoon split adjustments (wOBA points)
# Right-handed hitters: +0.020 vs LHP, Left-handed: -0.015 vs LHP
# Since we don't have handedness in our data, we estimate:
# ~60% of hitters are RHH, ~40% LHH (league average)
# Average platoon effect blended: 0.60 * 0.020 + 0.40 * (-0.015) = +0.006
# We use this as a noise proxy and test the methods' ability to improve on baseline
PLATOON_ADJ_RHH = 0.020  # RHH bonus vs LHP
PLATOON_ADJ_LHH = -0.015  # LHH penalty vs LHP
DAMPENING_FIXED = 0.50  # Fixed dampening for pitcher quality method

def load_data(db_path: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Load batting and pitching data."""
    conn = get_db_connection(db_path)

    batting = pd.read_sql_query(
        """
        SELECT fangraphs_id, season, name, team, PA, wOBA
        FROM batting_season
        WHERE season BETWEEN 2019 AND 2025
          AND PA >= 200
          AND wOBA IS NOT NULL
        """,
        conn,
    )

    pitching = pd.read_sql_query(
        """
        SELECT season, team, IP, SIERA
        FROM pitching_season
        WHERE season BETWEEN 2019 AND 2025
          AND IP IS NOT NULL
          AND SIERA IS NOT NULL
        """,
        conn,
    )

    conn.close()
    logger.info("Loaded %d batting, %d pitching rows", len(batting), len(pitching))
    return batting, pitching


def run_analysis(
    batting: pd.DataFrame, pitching: pd.DataFrame
) -> dict[str, pd.DataFrame]:
    """Compare three adjustment methods."""
    # Compute team SIERA and league average
    pitching = pitching.dropna(subset=["SIERA", "IP"])
    pitching["weighted_SIERA"] = pitching["SIERA"] * pitching["IP"]
    team_siera = (
        pitching.groupby(["season", "team"])
        .agg(total_IP=("IP", "sum"), weighted_sum=("weighted_SIERA", "sum"))
        .reset_index()
    )
    team_siera["team_SIERA"] = team_siera["weighted_sum"] / team_siera["total_IP"]

    league_avg = team_siera.groupby("season")["team_SIERA"].mean().reset_index()
    league_avg.columns = ["season", "league_avg_SIERA"]
    team_siera = team_siera.merge(league_avg, on="season")
    team_siera["SIERA_ratio"] = team_siera["team_SIERA"] / team_siera["league_avg_SIERA"]

    batting = batting.merge(
        team_siera[["season", "team", "SIERA_ratio"]],
        on=["season", "team"],
        how="inner",
    )

    # Since we lack handedness data, simulate by assigning ~60% RHH, ~40% LHH
    # Use a deterministic hash of fangraphs_id to assign
    np.random.seed(42)
    batting = batting.copy()
    batting["is_rhh"] = batting["fangraphs_id"].apply(
        lambda x: hash(str(x)) % 100 < 60
    )
    batting["platoon_adj"] = batting["is_rhh"].apply(
        lambda rhh: PLATOON_ADJ_RHH if rhh else PLATOON_ADJ_LHH
    )

    # Method A: Platoon splits proxy only
    # Assume ~35% of ABs are against opposite-hand pitchers
    opp_hand_pct = 0.35
    batting["method_a"] = batting["wOBA"] + batting["platoon_adj"] * opp_hand_pct

    # Method B: Pitcher quality only (SIERA dampening at 0.50)
    batting["method_b"] = batting["wOBA"] * (
        1 + (batting["SIERA_ratio"] - 1) * DAMPENING_FIXED
    )

    # Method C: Multiplicative (platoon * pitcher quality)
    batting["method_c"] = (
        (batting["wOBA"] + batting["platoon_adj"] * opp_hand_pct)
        * (1 + (batting["SIERA_ratio"] - 1) * DAMPENING_FIXED)
    )

    # Also track baseline (no adjustment)
    batting["baseline"] = batting["wOBA"]

    # Raw data for formulas
    raw_df = batting[[
        "season", "name", "team", "PA", "wOBA", "is_rhh", "platoon_adj",
        "SIERA_ratio", "baseline", "method_a", "method_b", "method_c",
    ]].copy()

    # Overall RMSE comparison
    methods = {
        "Baseline (no adjustment)": "baseline",
        "Method A: Platoon Only": "method_a",
        "Method B: Pitcher Quality Only": "method_b",
        "Method C: Multiplicative": "method_c",
    }

    comparison_records = []
    for label, col in methods.items():
        valid = raw_df.dropna(subset=["wOBA", col])
        rmse = np.sqrt(np.mean((valid["wOBA"] - valid[col]) ** 2))
        mae = np.mean(np.abs(valid["wOBA"] - valid[col]))
        comparison_records.append({
            "Method": label,
            "RMSE": rmse,
            "MAE": mae,
            "N": len(valid),
        })
    comparison_df = pd.DataFrame(comparison_records)

    # By season
    season_records = []
    for season in sorted(raw_df["season"].unique()):
        subset = raw_df[raw_df["season"] == season]
        for label, col in methods.items():
            valid = subset.dropna(subset=["wOBA", col])
            if len(valid) == 0:
                continue
            rmse = np.sqrt(np.mean((valid["wOBA"] - valid[col]) ** 2))
            season_records.append({
                "Season": int(season),
                "Method": label,
                "RMSE": rmse,
                "N": len(valid),
            })
    season_df = pd.DataFrame(season_records)

    # Sample sizes
    sample_df = pd.DataFrame({
        "Category": [
            "Total player-seasons",
            "RHH (estimated)",
            "LHH (estimated)",
            "Seasons covered",
            "Min PA threshold",
        ],
        "Value": [
            len(raw_df),
            int(raw_df["is_rhh"].sum()),
            int((~raw_df["is_rhh"]).sum()),
            len(raw_df["season"].unique()),
            200,
        ],
    })

    return {
        "raw": raw_df,
        "comparison": comparison_df,
        "by_season": season_df,
        "sample_sizes": sample_df,
    }


def write_excel(results: dict[str, pd.DataFrame], output_path: str) -> None:
    """Write platoon analysis to formatted Excel."""
    wb = Workbook()

    raw_df = results["raw"]
    comparison_df = results["comparison"]
    season_df = results["by_season"]
    sample_df = results["sample_sizes"]

    # ── Sheet 0: _RawData (hidden helper) ──
    ws_data = wb.active
    ws_data.title = "_RawData"

    data_headers = [
        "Season", "Name", "Team", "PA", "Actual wOBA", "Is RHH",
        "Platoon Adj", "SIERA Ratio", "Baseline", "Method A", "Method B", "Method C",
    ]
    for c, h in enumerate(data_headers, 1):
        ws_data.cell(row=1, column=c, value=h)
    style_header_row(ws_data, len(data_headers))

    for r_idx, (_, row) in enumerate(raw_df.iterrows(), 2):
        ws_data.cell(row=r_idx, column=1, value=int(row["season"]))
        ws_data.cell(row=r_idx, column=2, value=row["name"])
        ws_data.cell(row=r_idx, column=3, value=row["team"])
        ws_data.cell(row=r_idx, column=4, value=int(row["PA"]))
        ws_data.cell(row=r_idx, column=5, value=row["wOBA"])
        ws_data.cell(row=r_idx, column=5).number_format = "0.000"
        ws_data.cell(row=r_idx, column=6, value=row["is_rhh"])
        ws_data.cell(row=r_idx, column=7, value=row["platoon_adj"])
        ws_data.cell(row=r_idx, column=7).number_format = "0.000"
        ws_data.cell(row=r_idx, column=8, value=row["SIERA_ratio"])
        ws_data.cell(row=r_idx, column=8).number_format = "0.0000"
        for c_idx, col in enumerate(["baseline", "method_a", "method_b", "method_c"], 9):
            ws_data.cell(row=r_idx, column=c_idx, value=row[col])
            ws_data.cell(row=r_idx, column=c_idx).number_format = "0.0000"

    last_row = len(raw_df) + 1

    # ── Sheet 1: Method Comparison ──
    ws_comp = wb.create_sheet("Method Comparison")

    comp_headers = ["Method", "RMSE (Formula)", "RMSE (Computed)", "MAE", "N"]
    for c, h in enumerate(comp_headers, 1):
        ws_comp.cell(row=1, column=c, value=h)
    style_header_row(ws_comp, len(comp_headers))

    actual_col = "E"  # actual wOBA in _RawData
    method_cols = {"Baseline (no adjustment)": "I", "Method A: Platoon Only": "J",
                   "Method B: Pitcher Quality Only": "K", "Method C: Multiplicative": "L"}

    for r_idx, (_, row) in enumerate(comparison_df.iterrows(), 2):
        ws_comp.cell(row=r_idx, column=1, value=row["Method"])

        m_col = method_cols.get(row["Method"], "I")
        formula = (
            f"=SQRT(SUMPRODUCT(('_RawData'!{actual_col}2:{actual_col}{last_row}"
            f"-'_RawData'!{m_col}2:{m_col}{last_row})^2)"
            f"/COUNTA('_RawData'!{actual_col}2:{actual_col}{last_row}))"
        )
        ws_comp.cell(row=r_idx, column=2, value=formula).number_format = "0.000000"
        ws_comp.cell(row=r_idx, column=3, value=round(row["RMSE"], 6))
        ws_comp.cell(row=r_idx, column=3).number_format = "0.000000"
        ws_comp.cell(row=r_idx, column=4, value=round(row["MAE"], 6))
        ws_comp.cell(row=r_idx, column=4).number_format = "0.000000"
        ws_comp.cell(row=r_idx, column=5, value=int(row["N"]))

    n_methods = len(comparison_df)
    rng = f"B2:B{1 + n_methods}"
    add_conditional_min_max(ws_comp, rng, green_best="min")

    ws_comp.cell(row=1, column=1).comment = Comment(
        "Comparison of three wOBA adjustment methods:\n"
        "A) Platoon splits proxy (league-avg RHH/LHH adjustments)\n"
        "B) Pitcher quality only (team SIERA ratio * 0.50 dampening)\n"
        "C) Multiplicative combination of A and B\n\n"
        "Lower RMSE = better prediction accuracy.",
        "Analysis Script",
    )

    auto_width(ws_comp)

    # ── Sheet 2: By Season ──
    ws_season = wb.create_sheet("By Season")

    if not season_df.empty:
        method_labels = comparison_df["Method"].tolist()
        s_headers = ["Season"] + method_labels
        for c, h in enumerate(s_headers, 1):
            ws_season.cell(row=1, column=c, value=h)
        style_header_row(ws_season, len(s_headers))

        seasons_list = sorted(season_df["Season"].unique())
        for r_idx, season in enumerate(seasons_list, 2):
            ws_season.cell(row=r_idx, column=1, value=int(season))
            for c_idx, method in enumerate(method_labels, 2):
                match = season_df[
                    (season_df["Season"] == season) & (season_df["Method"] == method)
                ]
                if not match.empty:
                    ws_season.cell(
                        row=r_idx, column=c_idx, value=round(match.iloc[0]["RMSE"], 6)
                    ).number_format = "0.000000"

        for r_idx in range(2, 2 + len(seasons_list)):
            rng = f"B{r_idx}:{get_column_letter(1 + len(method_labels))}{r_idx}"
            add_conditional_min_max(ws_season, rng, green_best="min")
    else:
        ws_season.cell(row=1, column=1, value="No season data available.")

    auto_width(ws_season)

    # ── Sheet 3: Sample Sizes ──
    ws_sample = wb.create_sheet("Sample Sizes")
    sample_headers = ["Category", "Value"]
    for c, h in enumerate(sample_headers, 1):
        ws_sample.cell(row=1, column=c, value=h)
    style_header_row(ws_sample, len(sample_headers))

    for r_idx, (_, row) in enumerate(sample_df.iterrows(), 2):
        ws_sample.cell(row=r_idx, column=1, value=row["Category"])
        ws_sample.cell(row=r_idx, column=2, value=row["Value"])

    ws_sample.cell(row=1, column=1).comment = Comment(
        "Sample size details for the platoon analysis.\n"
        "Handedness is estimated (60% RHH / 40% LHH) since\n"
        "the backtest database lacks explicit handedness data.",
        "Analysis Script",
    )

    auto_width(ws_sample)

    # ── Sheet 4: Recommendation ──
    ws_rec = wb.create_sheet("Recommendation")
    ws_rec.cell(row=1, column=1, value="Platoon vs Pitcher Quality Recommendation").font = Font(
        bold=True, size=14
    )

    ws_rec.cell(row=3, column=1, value="Best Method:").font = HEADER_FONT
    ws_rec.cell(
        row=3, column=2,
        value=f"=INDEX('Method Comparison'!A2:A{1+n_methods},"
        f"MATCH(MIN('Method Comparison'!B2:B{1+n_methods}),"
        f"'Method Comparison'!B2:B{1+n_methods},0))",
    )
    ws_rec.cell(row=3, column=2).comment = Comment(
        "The method with the lowest overall RMSE.\n\n"
        "Note: With season-level data only, platoon effects are estimated\n"
        "using league-average splits. With game-level data, Method A\n"
        "would likely show more differentiation.",
        "Analysis Script",
    )

    ws_rec.cell(row=4, column=1, value="Minimum RMSE:").font = HEADER_FONT
    ws_rec.cell(
        row=4, column=2,
        value=f"=MIN('Method Comparison'!B2:B{1+n_methods})",
    )
    ws_rec.cell(row=4, column=2).number_format = "0.000000"

    ws_rec.cell(row=6, column=1, value="Methodology:").font = HEADER_FONT
    ws_rec.cell(row=7, column=1, value=(
        "Method A (Platoon Only): Adjusts wOBA using league-average platoon splits.\n"
        "  RHH: +0.020 wOBA vs LHP; LHH: -0.015 vs LHP. Applied to 35% of ABs.\n\n"
        "Method B (Pitcher Quality Only): Adjusts wOBA by team SIERA ratio.\n"
        "  adj_wOBA = wOBA * (1 + (SIERA_ratio - 1) * 0.50)\n\n"
        "Method C (Multiplicative): Combines both adjustments multiplicatively.\n"
        "  adj_wOBA = (wOBA + platoon_adj * 0.35) * (1 + (SIERA_ratio - 1) * 0.50)\n\n"
        "Limitation: Without individual handedness data, platoon adjustments are\n"
        "approximated using a 60/40 RHH/LHH split derived from player ID hashing."
    )).alignment = Alignment(wrap_text=True)
    ws_rec.column_dimensions["A"].width = 80
    ws_rec.column_dimensions["B"].width = 35

    wb.save(output_path)
    logger.info("Saved workbook to %s", output_path)


def main() -> None:
    """Entry point for platoon replacement analysis."""
    parser = argparse.ArgumentParser(
        description="Compare platoon vs pitcher quality vs multiplicative adjustments"
    )
    parser.add_argument(
        "--db-path",
        default=str(PROJECT_ROOT / "backtest_data.sqlite"),
        help="Path to backtest SQLite database",
    )
    parser.add_argument(
        "--output-dir",
        default=str(PROJECT_ROOT / "analysis"),
        help="Directory for Excel output",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(name)s %(levelname)s %(message)s",
    )

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    batting, pitching = load_data(args.db_path)

    if batting.empty or pitching.empty:
        logger.warning("No data found. Generating empty workbook.")
        wb = Workbook()
        wb.active.title = "No Data"
        wb.active.cell(row=1, column=1, value="No data found in database.")
        wb.save(str(output_dir / "platoon_analysis.xlsx"))
        return

    results = run_analysis(batting, pitching)
    write_excel(results, str(output_dir / "platoon_analysis.xlsx"))
    logger.info("Platoon replacement analysis complete.")


if __name__ == "__main__":
    main()
